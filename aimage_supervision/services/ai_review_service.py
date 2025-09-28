import asyncio
import io
import json
import os
import tempfile
import time
import traceback
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from uuid import UUID

import boto3
import httpx
import pandas as pd
from google import genai
from google.genai import types
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font
from PIL import Image
from pydantic import BaseModel, Field
from tortoise.exceptions import DoesNotExist
from tortoise.expressions import Q
from tortoise.transactions import in_transaction

from aimage_supervision.clients.aws_s3 import (
    download_file_content_from_s3_sync, download_image_from_s3,
    download_video_from_s3)
from aimage_supervision.endpoints.bounding_box_detect import \
    detect_bounding_boxes_new_sync
from aimage_supervision.enums import (AiReviewMode, AiReviewProcessingStatus,
                                      SubtaskType)
from aimage_supervision.models import (AiReview, AiReviewFindingEntry,
                                       Character, ReviewPointDefinition,
                                       ReviewPointDefinitionVersion, Subtask,
                                       User)
from aimage_supervision.prompts.review_point_prompts import (
    copyright_review_prompt, get_speaker_prompt, get_target_prompt,
    ng_review_prompt, visual_review_prompt)
from aimage_supervision.schemas import AiDetectedElement, AiDetectedElements
from aimage_supervision.schemas import AiReview as AiReviewSchema
from aimage_supervision.schemas import \
    AiReviewFindingEntry as AiReviewFindingEntrySchema
from aimage_supervision.schemas import (AiReviewFindingEntryInDB, AiReviewInDB,
                                        ReviewPointDefinitionVersionInDB)
from aimage_supervision.services.ai_review_pipeline import (
    BoundingBoxes, _generate_findings_for_copyright_review_sync,
    _generate_findings_for_ng_review_sync,
    _generate_findings_for_visual_review_sync,
    convert_bounding_boxes_to_x_y_width_height)
from aimage_supervision.services.ai_review_pipeline_multi_agent import \
    generate_findings_for_visual_review_multi_agent_sync
from aimage_supervision.services.rpd_create_service import split_review_prompt
from aimage_supervision.services.rpd_filter import filter_rpd_async
from aimage_supervision.settings import MAX_CONCURRENT, logger
from aimage_supervision.utils.video_utils import (
    detect_video_scenes_and_extract_frames, process_image_or_video_bytes)

# Model constants
QUALITY_MODEL = "gemini-2.5-pro"
SPEED_MODEL = "gemini-2.5-flash"

# Configure file handler for debugging in threads


if os.getenv('GEMINI_API_KEY'):
    gemini_client = genai.Client(api_key=os.getenv('GEMINI_API_KEY'))
else:
    gemini_client = None
    logger.warning(
        "GEMINI_API_KEY not set. detect_elements will not function.")


async def get_active_review_point_versions(
    project_id: Optional[UUID] = None,
    cr_check: Optional[bool] = False,
    rpd_ids: Optional[List[UUID]] = None
) -> List[ReviewPointDefinitionVersionInDB]:
    """
    Fetches active ReviewPointDefinitionVersion records.
    - If cr_check is True, fetches only the copyright review.
    - If rpd_ids are provided, fetches only those specific RPDs.
    - If neither is provided, returns an empty list.
    """
    # Keep the cr_check logic for backward compatibility. It takes precedence.
    if cr_check:
        query = ReviewPointDefinitionVersion.filter(
            is_active_version=True,
            review_point_definition__is_active=True,
            review_point_definition__key='copyright_review'
        )
        if project_id:
            query = query.filter(
                review_point_definition__project_id=project_id)

        active_versions = await query.prefetch_related('review_point_definition').all()
        return [ReviewPointDefinitionVersionInDB.from_orm(v) for v in active_versions]

    # New logic for rpd_ids
    if not rpd_ids:
        # Return empty list if no IDs are provided (and cr_check is false)
        return []

    query = ReviewPointDefinitionVersion.filter(
        is_active_version=True,
        review_point_definition__is_active=True,
        review_point_definition_id__in=rpd_ids  # Filter by the provided IDs
    )

    if project_id:
        query = query.filter(review_point_definition__project_id=project_id)

    active_versions = await query.prefetch_related('review_point_definition').all()

    return [ReviewPointDefinitionVersionInDB.from_orm(v) for v in active_versions]


async def _generate_single_rpd_findings(
    image_bytes: bytes,
    image_width: int,
    image_height: int,
    text: str,
    rpd_version: ReviewPointDefinitionVersionInDB,
    semaphore: asyncio.Semaphore,
    model_name: str
) -> List[Dict[str, Any]]:
    """
    为单个RPD版本生成findings数据，使用asyncio.to_thread实现真正的并行
    """
    # 检查该rpd是否ready，如果没有ready则轮询等待
    if not rpd_version.is_ready_for_ai_review:
        print(f"RPD版本 {rpd_version.title} 尚未准备就绪，开始轮询等待...")

        # 轮询等待RPD准备完成
        polling_start_time = time.time()
        max_wait_time = 5 * 60  # 5分钟
        polling_interval = 30   # 30秒

        while True:
            # 检查是否超时
            elapsed_time = time.time() - polling_start_time
            if elapsed_time >= max_wait_time:
                print(
                    f"RPD版本 {rpd_version.title} 等待超时({max_wait_time}秒)，返回空结果")
                # 返回固定的空结果
                return [{
                    'rpd_version': rpd_version,
                    'finding_dict': {
                        'description': 'RPD前処理失敗しました',
                        'severity': 'alert',
                        'suggestion': 'しばらく時間をおいてから再試行してください',
                        'tag': 'system_timeout'
                    }
                }]

            # 重新从数据库获取最新状态
            try:
                fresh_rpd = await ReviewPointDefinitionVersion.get(id=rpd_version.id).prefetch_related('review_point_definition')
                if fresh_rpd.is_ready_for_ai_review:
                    print(f"RPD版本 {rpd_version.title} 现在已准备就绪，继续处理...")
                    # 更新rpd_version对象为最新状态，保持原有的parent_key
                    original_parent_key = rpd_version.parent_key
                    rpd_version = ReviewPointDefinitionVersionInDB.from_orm(
                        fresh_rpd)
                    rpd_version.parent_key = original_parent_key
                    break
                else:
                    print(
                        f"RPD版本 {rpd_version.title} 仍未就绪，等待{polling_interval}秒后重试...")
                    await asyncio.sleep(polling_interval)
            except Exception as e:
                print(f"轮询RPD状态时出错: {e}")
                await asyncio.sleep(polling_interval)

    # 立即记录任务开始的时间戳
    task_start_time = time.time()
    print(f"*** 任务 {rpd_version.title},{text} 真正开始执行 at {task_start_time}")

    start_time = time.time()
    print(f"开始处理 rpd_version: {rpd_version.title} at {start_time}")

    findings_data = []

    try:
        if rpd_version.parent_key == 'general_ng_review':
            api_start_time = time.time()
            print(f"*** {rpd_version.title} 开始API调用 at {api_start_time}")

            # 使用 asyncio.to_thread 在线程中执行同步API调用
            findings_list = await asyncio.to_thread(
                _generate_findings_for_ng_review_sync, image_bytes, rpd_version, model_name
            )

            api_end_time = time.time()
            print(
                f"API调用完成 - {rpd_version.title} at {api_end_time} (耗时: {api_end_time - api_start_time:.2f}秒)")
            print(f"result: {findings_list}")

            for finding_dict in findings_list:
                findings_data.append({
                    'rpd_version': rpd_version,
                    'finding_dict': {
                        'description': finding_dict['description'],
                        'severity': finding_dict['severity'],
                        'suggestion': finding_dict['suggestion'],
                        'tag': finding_dict.get('tag', '')  # 确保包含tag字段
                    }
                })
            if findings_list == []:
                findings_data.append({
                    'rpd_version': rpd_version,
                    'finding_dict': {
                        'description': '特に注意すべきものはないです。',
                        'severity': 'safe',
                        'suggestion': '',
                        'tag': ''
                    }
                })

        elif rpd_version.parent_key == 'copyright_review':
            api_start_time = time.time()
            print(f"*** {rpd_version.title} 开始API调用 at {api_start_time}")

            # 使用新的边界框检测API
            bounding_boxes = await asyncio.to_thread(
                detect_bounding_boxes_new_sync, image_bytes, text
            )
            print(f"bounding_boxes: {bounding_boxes}")
            # 选择置信度最高的bounding box
            if bounding_boxes.bounding_boxes:
                highest_confidence_box = max(
                    bounding_boxes.bounding_boxes, key=lambda box: box.confidence)
                bounding_boxes = BoundingBoxes(
                    bounding_boxes=[highest_confidence_box])
            description, severity, suggestion = await asyncio.to_thread(
                _generate_findings_for_copyright_review_sync, image_bytes, bounding_boxes, rpd_version, model_name
            )

            api_end_time = time.time()
            print(
                f"API调用完成 - {rpd_version.title} at {api_end_time} (耗时: {api_end_time - api_start_time:.2f}秒)")
            bbox = convert_bounding_boxes_to_x_y_width_height(
                bounding_boxes, image_width, image_height)
            if description and bbox != []:
                findings_data.append({
                    'rpd_version': rpd_version,
                    'finding_dict': {
                        'description': description,
                        'severity': severity,
                        'suggestion': suggestion
                    },
                    'bounding_boxes': bbox[0]
                })
            elif description:
                findings_data.append({
                    'rpd_version': rpd_version,
                    'finding_dict': {
                        'description': description,
                        'severity': severity,
                        'suggestion': suggestion
                    },
                    'bounding_boxes': {
                        'x': 0,
                        'y': 0,
                        'width': 0,
                        'height': 0
                    }
                })
            else:
                findings_data.append({
                    'rpd_version': rpd_version,
                    'finding_dict': {
                        'description': '特に注意すべきものはないです。',
                        'severity': 'safe',
                        'suggestion': ''
                    },
                    'bounding_boxes': {
                        'x': 0,
                        'y': 0,
                        'width': 0,
                        'height': 0
                    }
                })

        elif rpd_version.parent_key == 'visual_review':
            if rpd_version.rpd_type != "classification tasks" and rpd_version.rpd_type != "right/wrong tasks":
                print("Unsupported RPD type for visual review:",
                      rpd_version.rpd_type)
                findings_data.append({
                    'rpd_version': rpd_version,
                    'finding_dict': {
                        'description': rpd_version.title,
                        'severity': 'alert',
                        'suggestion': "以下の内容をチェックしてください：" + rpd_version.user_instruction if rpd_version.user_instruction else "以下の内容をチェックしてください："
                    }
                })
            else:
                api_start_time = time.time()
                print(f"*** {rpd_version.title} 开始API调用 at {api_start_time}")

                # 使用新的边界框检测API
                bounding_boxes = await asyncio.to_thread(
                    detect_bounding_boxes_new_sync, image_bytes, text
                )
                print(f"bounding_boxes: {bounding_boxes}")
                # 选择置信度最高的bounding box
                if bounding_boxes.bounding_boxes:
                    highest_confidence_box = max(
                        bounding_boxes.bounding_boxes, key=lambda box: box.confidence)
                    bounding_boxes = BoundingBoxes(
                        bounding_boxes=[highest_confidence_box])
                description, severity, suggestion = await asyncio.to_thread(
                    generate_findings_for_visual_review_multi_agent_sync, image_bytes, rpd_version, bounding_boxes, model_name
                )

                api_end_time = time.time()
                print(
                    f"API调用完成 - {rpd_version.title} at {api_end_time} (耗时: {api_end_time - api_start_time:.2f}秒)")

                bbox = convert_bounding_boxes_to_x_y_width_height(
                    bounding_boxes, image_width, image_height)
                if description and bbox != []:
                    findings_data.append({
                        'rpd_version': rpd_version,
                        'finding_dict': {
                            'description': description,
                            'severity': severity,
                            'suggestion': suggestion
                        },
                        'bounding_boxes': bbox[0]
                    })
                elif description:
                    findings_data.append({
                        'rpd_version': rpd_version,
                        'finding_dict': {
                            'description': description,
                            'severity': severity,
                            'suggestion': suggestion
                        },
                        'bounding_boxes': {
                            'x': 0,
                            'y': 0,
                            'width': 0,
                            'height': 0
                        }
                    })
                else:
                    findings_data.append({
                        'rpd_version': rpd_version,
                        'finding_dict': {
                            'description': '特に注意すべきものはないです。',
                            'severity': 'safe',
                            'suggestion': ''
                        },
                        'bounding_boxes': {
                            'x': 0,
                            'y': 0,
                            'width': 0,
                            'height': 0
                        }
                    })

    except Exception as e:
        print(f"处理 RPD 版本 {rpd_version.title} 时出错: {e}")
        # 返回空列表而不是抛出异常

    end_time = time.time()
    print(
        f"*** 完成处理 rpd_version: {rpd_version.title} at {end_time} (总耗时: {end_time - task_start_time:.2f}秒)")
    return findings_data


async def _generate_findings_for_all_rpd_versions(
    image_bytes: bytes,
    image_width: int,
    image_height: int,
    content_type: SubtaskType,
    content_metadata: Dict[str, Any],
    active_rpd_versions: List[ReviewPointDefinitionVersionInDB],
    ai_review_orm: AiReview,
    mode: AiReviewMode
) -> List[AiReviewFindingEntry]:
    """
    并行为所有RPD版本生成findings，然后批量保存到数据库
    """
    start_time = time.time()

    model_name = QUALITY_MODEL if mode == AiReviewMode.QUALITY else SPEED_MODEL
    logger.info(
        f"Using model: {model_name} for AI review in mode: {mode.value}")

    # 设置更大的并发数，因为现在每个任务都有独立的客户端
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)

    print(f"开始处理 {len(active_rpd_versions)} 个RPD版本")

    # 检查中断信号
    if await _check_cancellation_signal(ai_review_orm):
        print("检测到中断信号，停止处理")
        return []

    # RPD 过滤阶段
    # filtered_rpd_versions = active_rpd_versions
    filtered_rpd_versions = []

    try:
        print(f"开始 RPD 过滤，原始 RPD 数量: {len(active_rpd_versions)}")

        # 检查中断信号
        if await _check_cancellation_signal(ai_review_orm):
            print("在 RPD 过滤前检测到中断信号，停止处理")
            return []

        # 执行 RPD 过滤 - 直接使用 ReviewPointDefinitionVersionInDB，无需转换
        filter_start_time = time.time()

        # 创建取消检查回调函数
        async def cancellation_check():
            return await _check_cancellation_signal(ai_review_orm)

        filtered_rpd_versions_raw = await filter_rpd_async(
            image_bytes,
            active_rpd_versions,  # 直接使用，无需转换
            max_concurrent=MAX_CONCURRENT,  # 与并行处理保持一致的并发数
            cancellation_check_callback=cancellation_check
        )
        filter_end_time = time.time()
        print(f"RPD 过滤完成，耗时: {filter_end_time - filter_start_time:.2f}秒")
        print(f"过滤后 RPD 数量: {len(filtered_rpd_versions_raw)}")

        # 检查中断信号
        if await _check_cancellation_signal(ai_review_orm):
            print("在 RPD 过滤后检测到中断信号，停止处理")
            return []

        # 直接使用过滤后的结果，无需转换
        filtered_rpd_versions = filtered_rpd_versions_raw

        print(f"最终用于处理的 RPD 数量: {len(filtered_rpd_versions)}")

    except Exception as e:
        print(f"RPD 过滤过程中出错: {e}")
        print("将使用原始 RPD 列表继续处理")
        # filtered_rpd_versions = active_rpd_versions
        for rpd_version in active_rpd_versions:
            filtered_rpd_versions.append({
                'rpd_version': rpd_version,
                'word': rpd_version.title
            })

    # 如果过滤后没有 RPD，直接返回空列表
    if not filtered_rpd_versions:
        print("过滤后没有有效的 RPD，跳过处理")
        return []

    print(f"开始并行处理 {len(filtered_rpd_versions)} 个过滤后的RPD版本")

    # 创建任务 - 立即启动所有任务
    print("创建并立即启动所有任务...")
    tasks = []
    for i, rpd_version_dict in enumerate(filtered_rpd_versions):
        rpd_version = rpd_version_dict['rpd_version']
        text = rpd_version_dict['word']
        print(f"为 {rpd_version.title} 创建任务 {i} at {time.time()}")
        # 使用 asyncio.create_task 立即启动任务
        task = asyncio.create_task(
            _generate_single_rpd_findings(
                image_bytes,  image_width, image_height, text, rpd_version, semaphore, model_name),
            name=f"rpd_task_{rpd_version.title}"
        )
        tasks.append(task)
        print(f"任务 {rpd_version.title} 已创建并启动 at {time.time()}")

    print(f"所有 {len(tasks)} 个任务已创建并启动，开始等待完成...")

    # 使用可中断的等待机制
    done = await _wait_for_tasks_with_cancellation(tasks, ai_review_orm, 2.0)

    # 如果返回空列表，说明被中断了
    if not done:
        return []

    parallel_time = time.time()
    print(f"并行处理完成，耗时: {parallel_time - start_time:.2f}秒")

    # 处理结果并创建数据库对象
    all_findings = []
    for task in done:
        try:
            findings_data_list = await task
            if isinstance(findings_data_list, Exception):
                print(f"任务执行异常: {findings_data_list}")
                continue

            for finding_data in findings_data_list:
                rpd_version = finding_data['rpd_version']
                finding_dict = finding_data['finding_dict']
                if 'bounding_boxes' in finding_data:
                    area = finding_data['bounding_boxes']
                else:
                    area = {'x': 0, 'y': 0, 'width': 0, 'height': 0}
                # 创建完整的AiReviewFindingEntry对象
                finding_entry = await AiReviewFindingEntry.create(
                    ai_review=ai_review_orm,
                    review_point_definition_version_id=rpd_version.id,
                    description=finding_dict['description'],
                    severity=finding_dict['severity'],
                    suggestion=finding_dict.get('suggestion', ''),
                    area=area,  # 默认区域
                    is_ai_generated=True,
                    status='pending_human_review',
                    tag=finding_dict.get('tag', ''),
                    reference_images=rpd_version.reference_images,
                    reference_source=rpd_version.title,
                    content_type=content_type,
                    content_metadata=content_metadata
                )
                all_findings.append(finding_entry)
        except Exception as e:
            print(f"处理任务结果时出错: {e}")

    db_end_time = time.time()
    print(
        f"数据库操作耗时: {db_end_time - parallel_time:.2f}秒, 总耗时: {db_end_time - start_time:.2f}秒")

    return all_findings


async def _process_single_video_frame(
    video_frame_info: Dict[str, Any],
    video_width: int,
    video_height: int,
    content_type: SubtaskType,
    active_rpd_versions: List[ReviewPointDefinitionVersionInDB],
    ai_review_orm: AiReview,
    mode: AiReviewMode,
    semaphore: asyncio.Semaphore
) -> List[AiReviewFindingEntry]:
    """
    处理单个视频帧，为其生成所有RPD的findings
    """
    async with semaphore:
        try:
            video_frame_bytes = video_frame_info["middle_frame_bytes"]
            content_metadata = {
                "scene_number": video_frame_info["scene_number"],
                "start_timestamp": video_frame_info["start_time"],
                "end_timestamp": video_frame_info["end_time"],
                "start_frame": video_frame_info["start_frame"],
                "end_frame": video_frame_info["end_frame"],
            }

            print(f"处理视频帧 - 场景 {content_metadata['scene_number']}")

            # 为这个视频帧生成所有RPD的findings
            findings = await _generate_findings_for_all_rpd_versions(
                video_frame_bytes, video_width, video_height, content_type,
                content_metadata, active_rpd_versions, ai_review_orm, mode
            )

            print(
                f"视频帧场景 {content_metadata['scene_number']} 处理完成，生成 {len(findings)} 个findings")
            return findings

        except Exception as e:
            print(f"处理视频帧时出错: {e}")
            return []


async def _generate_findings_for_video_frames_parallel(
    video_data: List[Dict[str, Any]],
    video_width: int,
    video_height: int,
    content_type: SubtaskType,
    active_rpd_versions: List[ReviewPointDefinitionVersionInDB],
    ai_review_orm: AiReview,
    mode: AiReviewMode,
    max_concurrent: Optional[int] = None
) -> List[AiReviewFindingEntry]:
    """
    并行处理所有视频帧，为每个帧生成AI审查findings
    """
    start_time = time.time()

    print(f"开始并行处理 {len(video_data)} 个视频帧")

    # 设置并发信号量
    if max_concurrent is None:
        max_concurrent = MAX_CONCURRENT
    semaphore = asyncio.Semaphore(max_concurrent)

    # 创建所有视频帧处理任务
    tasks = []
    for i, video_frame_info in enumerate(video_data):
        task = asyncio.create_task(
            _process_single_video_frame(
                video_frame_info, video_width, video_height, content_type,
                active_rpd_versions, ai_review_orm, mode, semaphore
            ),
            name=f"video_frame_task_{i}"
        )
        tasks.append(task)

    print(f"所有 {len(tasks)} 个视频帧任务已创建，开始等待完成...")

    # 使用可中断的等待机制
    done = await _wait_for_tasks_with_cancellation(tasks, ai_review_orm, 2.0)

    # 如果返回空列表，说明被中断了
    if not done:
        return []

    # 收集所有结果
    all_findings = []
    for task in done:
        try:
            findings = await task
            all_findings.extend(findings)
        except Exception as e:
            print(f"获取视频帧处理结果时出错: {e}")

    end_time = time.time()
    print(f"视频帧并行处理完成，总耗时: {end_time - start_time:.2f}秒")
    print(f"共生成 {len(all_findings)} 个findings")

    return all_findings

# ------------------------------------------------------------
# 称呼检查 (moved from check_speakers.py)
# ------------------------------------------------------------

# Speaker check related models (moved from check_speakers.py)


class SpeakResult(BaseModel):
    speaker: str = Field(description="The official name of the speaker")
    speaker_used_name: str = Field(
        description="The actual name used for the speaker in the conversation")


class TargetResult(BaseModel):
    target: List[str] = Field(
        description="List of official names of people mentioned in the conversation")
    target_used_names: List[str] = Field(
        description="List of actual names used in the conversation for each target")


class AliasCheckResult(BaseModel):
    wrong_list: List[str] = Field(description="List of wrong alias")
    correct_list: List[str] = Field(description="List of correct alias")


async def get_speaker(
    gemini_client: Any,
    cleaned_text: str,
    model_name: str,
    alias_dict: Dict[str, Dict[str, str]] = None,
    max_retries: int = 3,
    retry_delay: float = 1.0,
    timeout: float = 30.0,  # 添加超时参数,
) -> SpeakResult:
    """第一步：获取speaker"""
    # print(f"第一步：分析对话中的说话者: {cleaned_text}")

    # 获取所有可能的说话者
    official_speakers = sorted(list(alias_dict.keys()))

    for attempt in range(max_retries + 1):
        try:
            # print(f"🔄 第一步 API 调用尝试 {attempt + 1}/{max_retries + 1}")

            response = await asyncio.wait_for(
                asyncio.to_thread(
                    gemini_client.models.generate_content,
                    model=model_name,
                    contents=cleaned_text,
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json",
                        response_schema=SpeakResult,
                        system_instruction=get_speaker_prompt.replace(
                            "{json.dumps(official_speakers, ensure_ascii=False)}",
                            json.dumps(official_speakers, ensure_ascii=False)
                        ),
                        temperature=0.0,
                    )
                ),
                timeout=timeout
            )

            result = response.parsed
            return result

        except asyncio.TimeoutError:
            print(f"⏰ 第一步API调用超时 (尝试 {attempt + 1}), 超时时间: {timeout}秒")
            if attempt == max_retries:
                raise Exception(f"第一步API调用在 {max_retries + 1} 次尝试后仍然超时")

        except Exception as e:
            print(f"❌ 第一步API调用失败 (尝试 {attempt + 1}): {str(e)}")
            if attempt == max_retries:
                raise e

        if attempt < max_retries:
            print(f"⏳ 等待 {retry_delay} 秒后重试...")
            await asyncio.sleep(retry_delay)
            retry_delay *= 1.5

    return None


async def get_targets(
    gemini_client: Any,
    line: str,
    speaker: str,
    model_name: str,
    alias_dict: Dict[str, Dict[str, str]],
    all_characters: List[str],
    max_retries: int = 3,
    retry_delay: float = 1.0,
    timeout: float = 30.0,
    sheet_index: int = 0,
) -> TargetResult:
    """第二步：获取对话中提到的人物"""
    # print(f"🔍 第二步：分析对话中的目标人物(轮次: {sheet_index})")

    # 构造所有可能的称呼规则 (不含speaker)
    all_possible_aliases = set()
    for target, alias in alias_dict.get(speaker, {}).items():
        all_possible_aliases.add(f"{target} should be called as '{alias}'")

    possible_aliases_text = "".join(sorted(list(all_possible_aliases)))

    # 根据轮次动态添加指令

    for attempt in range(max_retries + 1):
        try:
            # print(f"🔄 第二步 API 调用尝试 {attempt + 1}/{max_retries + 1}")

            response = await asyncio.wait_for(
                asyncio.to_thread(
                    gemini_client.models.generate_content,
                    model=model_name,
                    contents=line,
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json",
                        response_schema=TargetResult,
                        system_instruction=get_target_prompt.replace(
                            "{speaker}", speaker
                        ).replace(
                            "{possible_aliases_text}", possible_aliases_text
                        ).replace(
                            "{json.dumps(all_characters, ensure_ascii=False)}",
                            json.dumps(all_characters, ensure_ascii=False)
                        ),
                        temperature=0.0,
                    )
                ),
                timeout=timeout
            )

            result = response.parsed
            return result, possible_aliases_text

        except asyncio.TimeoutError:
            print(f"⏰ 第二步API调用超时 (尝试 {attempt + 1}), 超时时间: {timeout}秒")
            if attempt == max_retries:
                raise Exception("第二步API调用超时")

        except Exception as e:
            print(f"❌ 第二步API调用失败 (尝试 {attempt + 1}): {str(e)}")
            if attempt == max_retries:
                raise e

        if attempt < max_retries:
            await asyncio.sleep(retry_delay)
            retry_delay *= 1.5
    return None, None


async def check_speaker(
    data: Dict[str, Any],
    model_name: str,
    max_retries: int = 3,
    retry_delay: float = 1.0,
    alias_dict: Dict[str, Dict[str, str]] = None,
    extra_info: List[Dict[str, Any]] = None,
    api_timeout: float = 30.0,  # 添加API超时参数
    sheet_index: int = 0
) -> Dict[str, Any]:
    api_start_time = time.time()

    # 如果 extra_info 未提供，则初始化为空列表
    if extra_info is None:
        extra_info = []

    try:
        gemini_client = genai.Client(api_key=os.environ["GEMINI_API_KEY"])

        # 预处理文本
        original_index = data['original_index']  # 使用正确的键名
        speaker = data['speaker']
        line = data['line']
        content = f"発話者:{speaker}, 会話内容:{line}"

        # print(f"📝 处理段落 {original_index}: {content}")

        # 第一步：获取speaker
        speaker_result = await get_speaker(
            gemini_client,
            content,
            model_name,
            alias_dict,
            max_retries,
            retry_delay,
            api_timeout,
        )

        # 第一轮结果
        first_result = {
            "original_index": original_index,  # 使用正确的键名
            "speaker": speaker_result.speaker.replace("鳳・ここな", "鳳·ここな"),
            "line": line,
            "speaker_used_name": speaker_result.speaker_used_name
        }

        # 获取所有可能的角色列表
        all_characters = list(alias_dict.keys())

        # 第二步：获取target
        target_result, possible_aliases_text = await get_targets(
            gemini_client,
            line,
            speaker_result.speaker,
            model_name,
            alias_dict,
            all_characters,
            max_retries,
            retry_delay,
            api_timeout,
            sheet_index=sheet_index,

        )

        second_result = {
            "original_index": original_index,  # 使用正确的键名
            "target": [target.replace("鳳・ここな", "鳳·ここな") for target in target_result.target],
            "target_used_names": target_result.target_used_names,
        }

        # 合并两轮结果
        combined_result = {
            **first_result,
            **second_result,
        }

        # 第三步: 机械检查别名使用 (新版三级检查逻辑)

        # 如果说话者不在称呼表中，则跳过检查，但标记为unknown_speaker状态
        speaker_name = combined_result["speaker"]
        if speaker_name not in alias_dict:
            # print(f"ℹ️ 说话者「{speaker_name}」不在称呼表 (alias_dict) 中，跳过机械检查。")
            return {
                **combined_result,
                "status": "unknown_speaker",
                "message": f"話者「{speaker_name}」は呼称表にありません。",
                "response_time": round(time.time() - api_start_time, 2),
                "error": None
            }

        found_errors = []
        found_warnings = []

        speaker_alias_rules = alias_dict.get(combined_result["speaker"], {})

        for i, target in enumerate(combined_result["target"]):
            if i >= len(combined_result["target_used_names"]):
                continue

            used_name = combined_result["target_used_names"][i]
            speaker_name = combined_result["speaker"]

            # 1. 常规规则检查
            correct_alias = speaker_alias_rules.get(target)
            if correct_alias and (used_name == correct_alias or used_name == target):
                # print(f"✅ (常规) 检查通过: {speaker_name} 对 {target} 使用 '{used_name}' 是正确的。")
                continue

            # 2. 特殊规则检查
            is_special_case = False
            if extra_info:
                for rule in extra_info:
                    if (rule.get('speaker') == speaker_name and
                        rule.get('target') == target and
                            rule.get('alias') == used_name):

                        conditions = rule.get('conditions', ['無特定条件'])
                        conditions_text = ', '.join(conditions) if isinstance(
                            conditions, list) else str(conditions)
                        message = f"[{speaker_name}] → [{target}] : 特殊な呼び方「{used_name}」が使われました。条件:{conditions_text}を確認してください。"
                        found_warnings.append(message)
                        # print(f"⚠️ (特殊) 检查通过: {speaker_name} 对 {target} 使用 '{used_name}' (条件: {conditions_text})")
                        is_special_case = True
                        break  # 找到一个匹配的特殊规则就够了

            if is_special_case:
                # todo 特殊規則対応不可
                continue

            # 3. 如果都不匹配，则为错误
            if correct_alias:
                message = f"{speaker_name} → {target} の呼び方は「{correct_alias}」ですが、「{used_name}」が使われました。"
            else:
                message = f"{speaker_name} → {target} の呼び方がありませんが、「{used_name}」が使われました。"

            found_errors.append(message)
            # print(f"❌ 发现错误: {message}")

        # 决定最终状态和消息
        # 状态分类：
        # - "ok": 完全正确，无问题
        # - "warning": 特殊规则情况，需要确认
        # - "error": 错误，需要修正
        final_status = "ok"
        final_message = f"セリフ：{speaker}: {line} - 問題なし"

        if found_errors:
            final_status = "error"
            # 错误信息优先，同时附带警告信息
            final_message = "\n".join(found_errors + found_warnings)
        elif found_warnings:
            # 如果只有警告，使用专门的特殊规则状态
            final_status = "warning"
            final_message = "\n".join(found_warnings)

        # 最终结果
        final_result = {
            **combined_result,
            "status": final_status,
            "message": final_message,
            "response_time": round(time.time() - api_start_time, 2),
            "error": None
        }

        return final_result

    except Exception as e:
        error_time = round(time.time() - api_start_time, 2)
        print(
            f"❌ 段落 {data.get('original_index', 'unknown')} 处理失败，耗时: {error_time}秒")
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {e}")
        print("详细追溯信息:")
        traceback.print_exc()
        return {
            "original_index": data.get('original_index', 'unknown'),  # 使用正确的键名
            "speaker": data.get('speaker', ''),
            "line": data.get('line', ''),
            "flag": False,
            "wrong_list": [],
            "correct_list": [],
            "response_time": error_time,
            "error": f"{type(e).__name__}: {e}{traceback.format_exc()}"
        }


def _load_appellation_from_rpds(text_rpds: List[ReviewPointDefinitionVersionInDB]) -> Dict[str, Dict[str, str]]:
    """
    从text_review RPD的reference_files中加载称呼表

    Args:
        text_rpds: text_review类型的RPD版本列表

    Returns:
        Dict[str, Dict[str, str]]: 称呼表数据
    """
    alias_dict = {}
    for rpd in text_rpds:
        if rpd.reference_files:
            # 查找JSON文件（称呼表）
            appellation_file = None
            for file_url in rpd.reference_files:
                if file_url.endswith('.json'):
                    appellation_file = file_url
                    break

            if appellation_file:
                print(f"从RPD {rpd.title} 加载称呼表: {appellation_file}")
                loaded_dict = _load_config_from_s3(appellation_file, "称呼表")
                if loaded_dict:
                    alias_dict.update(loaded_dict)
                    break  # 使用第一个成功加载的称呼表
            else:
                print(f"RPD {rpd.title} 的reference_files中未找到JSON文件")

    return alias_dict


def _load_config_from_s3(s3_path: str, description: str) -> Any:
    """从S3加载配置文件的通用函数"""
    try:
        from aimage_supervision.clients.aws_s3 import \
            download_file_content_from_s3_sync
        from aimage_supervision.settings import AWS_BUCKET_NAME

        print(f"正在从S3加载{description}: {s3_path}")

        # 解析S3路径，提取bucket名称和文件路径
        if s3_path.startswith('s3://'):
            # 完整S3 URL格式：s3://bucket-name/path/to/file
            parts = s3_path.replace('s3://', '').split('/', 1)
            if len(parts) == 2:
                bucket_name = parts[0]
                file_path = parts[1]
                print(f"解析S3 URL - Bucket: {bucket_name}, 文件路径: {file_path}")
            else:
                # 如果解析失败，使用默认bucket和原路径
                bucket_name = AWS_BUCKET_NAME
                file_path = s3_path
                print(f"S3 URL解析失败，使用默认bucket: {bucket_name}")
        else:
            # 相对路径格式，使用默认bucket
            bucket_name = AWS_BUCKET_NAME
            file_path = s3_path
            print(f"使用相对路径和默认bucket: {bucket_name}")

        # 使用解析出的bucket名称和文件路径下载文件
        file_content = download_file_content_from_s3_sync(
            file_path, bucket_name)
        content_str = file_content.decode('utf-8')
        config_data = json.loads(content_str)
        print(f"成功加载{description}，数据大小: {len(content_str)} 字符")
        return config_data
    except ValueError as e:
        # download_file_content_from_s3_sync 会在文件不存在时抛出 ValueError
        if "Error downloading file from S3" in str(e):
            print(f"S3文件不存在: {s3_path}")
        else:
            print(f"从S3加载{description}失败: {e}")
        return None
    except Exception as e:
        print(f"从S3加载{description}失败: {e}")
        return None


async def _process_single_text_part(
    prepared_data: Dict[str, Any],
    content_type: SubtaskType,
    text_rpds: List[ReviewPointDefinitionVersionInDB],
    ai_review_orm: AiReview,
    alias_dict: Dict[str, Any],
    semaphore: asyncio.Semaphore,
    model_name: str,
    extra_info: List[Dict[str, Any]] = None
) -> List[AiReviewFindingEntry]:
    """
    处理单个文本部分，为其生成所有RPD的findings（优化版本）
    整合了_load_alias_dict_once, _determine_severity, _create_findings_from_check_result的功能

    Args:
        prepared_data: 预处理的文本数据
        content_type: 内容类型
        text_rpds: 文本审查的RPD版本列表
        ai_review_orm: AI审查ORM对象
        alias_dict: 呼称表字典
        semaphore: 并发控制信号量
        mode: AI审查模式 (QUALITY/SPEED)，控制检查严格程度
        extra_info: 特殊规则列表，默认为空列表
                   格式: [{"speaker": "角色A", "target": "角色B", "alias": "特殊称呼", "conditions": ["条件1", "条件2"]}]
    """
    async with semaphore:
        try:
            # 如果没有呼称表或没有text_rpds，直接返回
            if not alias_dict or not text_rpds:
                return []

            # 执行说话者/称呼检查 - 直接使用预处理的数据
            check_result = await check_speaker(
                data=prepared_data,
                model_name=model_name,
                alias_dict=alias_dict,
                extra_info=extra_info or [],
                sheet_index=prepared_data["sheet_index"]
            )

            # 【整合_determine_severity功能】确定严重程度 - 与status统一
            # 映射关系（前端使用["risk", "alert", "safe"]）：
            # - status="ok" → severity="safe" (完全正确)
            # - status="warning" → severity="alert" (特殊规则，需确认)
            # - status="error" → severity="risk" (错误，需修正)
            # - status="unknown_speaker" → severity="alert" (speaker不在称呼表)
            # - status="unknown_target" → severity="alert" (target不在称呼表)
            status_val = check_result.get("status", "error")
            message_val = check_result.get("message", "") or ""

            if status_val == "error":
                severity = "risk"
            elif status_val == "warning":  # 特殊规则情况
                severity = "alert"
            elif status_val in ["unknown_speaker", "unknown_target"]:  # speaker或target不在称呼表
                severity = "alert"
            elif status_val == "ok":
                severity = "safe"
            else:
                severity = "risk"  # 未知状态默认为risk

            # 【整合_create_findings_from_check_result功能】根据检查结果创建findings
            messages = [m.strip()
                        for m in message_val.split("\n") if m.strip()]

            # 扩展content_metadata，包含check_speaker的完整结果
            content_metadata = {
                "original_index": prepared_data["original_index"],
                "sheet_index": prepared_data["sheet_index"],
                "speaker": prepared_data["speaker"],
                "line": prepared_data["line"],
                # 保存check_speaker的完整结果
                "check_speaker_result": {
                    "detected_speaker": check_result.get("speaker", prepared_data["speaker"]),
                    # 注意：check_speaker返回的是"target"字段
                    "target_list": check_result.get("target", []),
                    "target_used_names": check_result.get("target_used_names", []),
                    "speaker_used_name": check_result.get("speaker_used_name", ""),
                    "status": check_result.get("status", ""),
                }
            }

            created_findings = []
            for rpd_version in text_rpds:
                for msg in messages:
                    try:
                        finding_entry = await AiReviewFindingEntry.create(
                            ai_review=ai_review_orm,
                            review_point_definition_version_id=rpd_version.id,
                            description=msg,
                            severity=severity,
                            suggestion="",
                            area={"x": 0, "y": 0, "width": 0, "height": 0},
                            is_ai_generated=True,
                            status='pending_human_review',
                            tag='speaker_check',
                            reference_images=rpd_version.reference_images,
                            reference_source=rpd_version.title,
                            content_type=content_type,
                            content_metadata=content_metadata,
                        )
                        created_findings.append(finding_entry)
                    except Exception as create_err:
                        print(f"创建文本 finding 失败: {create_err}")

            print(
                f"文本部分 sheet {prepared_data.get('sheet_index', 'unknown')}, row {prepared_data.get('original_index', 'unknown')} 处理完成，生成 {len(created_findings)} 个findings"
            )
            return created_findings

        except NotImplementedError:
            print(
                f"文本处理功能尚未实现，跳过文本部分 {prepared_data.get('original_index', 'unknown')}")
            return []
        except Exception as e:
            print(f"处理文本部分时出错: {e}")
            return []


async def _generate_findings_for_text_parts_parallel(
    text_parts: List[Dict[str, Any]],
    content_type: SubtaskType,
    active_rpd_versions: List[ReviewPointDefinitionVersionInDB],
    ai_review_orm: AiReview,
    mode: AiReviewMode,
    local_excel_path: Optional[str] = None,
    original_s3_path: Optional[str] = None,
    max_concurrent: Optional[int] = None,
    extra_info: List[Dict[str, Any]] = None
) -> List[AiReviewFindingEntry]:
    """
    并行处理所有文本部分，为每个部分生成AI审查findings，并更新Excel文件（优化版本）

    Args:
        text_parts: 文本部分列表
        content_type: 内容类型
        active_rpd_versions: 活跃的RPD版本列表
        ai_review_orm: AI审查ORM对象
        mode: 审查模式
        local_excel_path: 本地Excel文件路径（可选）
        original_s3_path: 原始S3路径（可选）
        max_concurrent: 最大并发数
        extra_info: 特殊规则列表，默认为空列表
                   格式: [{"speaker": "角色A", "target": "角色B", "alias": "特殊称呼", "conditions": ["条件1", "条件2"]}]
    """
    start_time = time.time()
    model_name = QUALITY_MODEL if mode == AiReviewMode.QUALITY else SPEED_MODEL
    print(f"开始并行处理 {len(text_parts)} 个文本部分")
    print(f"使用模型: {model_name}")

    # 【优化1】提前过滤text_review类型的RPD，避免每个任务重复过滤
    text_rpds = [
        v for v in active_rpd_versions if v.parent_key == 'text_review']
    if not text_rpds:
        print("未找到 text_review 类型的RPD，跳过文本处理。")
        return []

    # 【优化2】从text_review RPD的reference_files中加载称呼表
    alias_dict = _load_appellation_from_rpds(text_rpds)
    if not alias_dict:
        print("未能从任何text_review RPD中加载称呼表，跳过文本处理。")
        return []

    # 设置并发信号量
    if max_concurrent is None:
        max_concurrent = MAX_CONCURRENT
    semaphore = asyncio.Semaphore(max_concurrent)

    tasks = []
    for i, text_part_data in enumerate(text_parts):
        prepared_data = {
            "original_index": text_part_data.get("original_index", 0),
            "speaker": text_part_data.get("speaker", ""),
            "line": text_part_data.get("line", ""),
            "sheet_index": text_part_data.get("sheet_index", 0)
        }

        task = asyncio.create_task(
            _process_single_text_part(
                prepared_data,           # 传递准备好的数据，而不是原始数据
                content_type,
                text_rpds,              # 传递过滤后的RPD列表
                ai_review_orm,
                alias_dict,             # 传递预加载的呼称表
                semaphore,
                model_name,
                extra_info              # 传递特殊规则
            ),
            name=f"text_part_task_{i}"
        )
        tasks.append(task)

    print(f"所有 {len(tasks)} 个文本部分任务已创建，开始等待完成...")

    # 使用可中断的等待机制
    done = await _wait_for_tasks_with_cancellation(tasks, ai_review_orm, 2.0)

    # 如果返回空列表，说明被中断了
    if not done:
        return []

    # 收集所有结果
    all_findings = []
    for task in done:
        try:
            findings = await task
            all_findings.extend(findings)
        except Exception as e:
            print(f"获取文本部分处理结果时出错: {e}")

    end_time = time.time()
    print(f"文本部分并行处理完成，总耗时: {end_time - start_time:.2f}秒")
    print(f"共生成 {len(all_findings)} 个findings")

    # 【新增功能】如果提供了Excel文件路径，则更新Excel文件并上传到S3
    if local_excel_path and original_s3_path and all_findings:
        try:
            print("开始更新Excel文件...")
            updated_s3_path = await _update_excel_with_findings(
                local_excel_path, text_parts, all_findings, original_s3_path, extra_info, text_rpds
            )
            print(f"Excel文件已更新并替换原文件: {updated_s3_path}")

            # 清理本地临时文件
            if os.path.exists(local_excel_path):
                os.unlink(local_excel_path)
                print(f"清理临时文件: {local_excel_path}")

        except Exception as e:
            print(f"更新Excel文件失败: {e}")
            # 即使Excel更新失败，也要清理临时文件
            if local_excel_path and os.path.exists(local_excel_path):
                os.unlink(local_excel_path)
    elif local_excel_path:
        # 如果没有findings或者参数不完整，也要清理临时文件
        try:
            if os.path.exists(local_excel_path):
                os.unlink(local_excel_path)
                print(f"清理临时文件: {local_excel_path}")
        except Exception as e:
            print(f"清理临时文件失败: {e}")

    return all_findings


def _analyze_target_used_names_status(
    check_speaker_result: Dict[str, Any],
    alias_dict: Dict[str, Any],
    extra_info: List[Dict[str, Any]] = None
) -> List[Dict[str, Any]]:
    """
    分析每个target_used_name的状态，返回详细的状态信息

    Args:
        check_speaker_result: check_speaker的结果
        alias_dict: 呼称表
        extra_info: 特殊规则列表

    Returns:
        List[Dict]: 每个target_used_name的状态信息
        [{"used_name": str, "status": "correct|error|special|unknown_speaker|unknown_target", "target": str}, ...]
    """
    if extra_info is None:
        extra_info = []

    results = []
    # 注意：根据check_speaker的实际返回结构，使用正确的字段名
    target_list = check_speaker_result.get(
        "target_list", [])  # 我们存储时用的是target_list
    target_used_names = check_speaker_result.get("target_used_names", [])
    speaker_name = check_speaker_result.get("detected_speaker", "")

    # 【新增】一致性检查：如果speaker不在称呼表中，标记所有target_used_names为unknown_speaker
    if speaker_name not in alias_dict:
        for i, target in enumerate(target_list):
            if i >= len(target_used_names):
                continue
            used_name = target_used_names[i]
            results.append({
                "used_name": used_name,
                "status": "unknown_speaker",
                "target": target,
                "reason": f"Speaker '{speaker_name}' not found in alias dictionary"
            })
        return results

    # 获取该说话者的称呼规则
    speaker_alias_rules = alias_dict.get(speaker_name, {})

    for i, target in enumerate(target_list):
        if i >= len(target_used_names):
            continue

        used_name = target_used_names[i]

        # 【新增】检查target是否在称呼表的所有角色列表中
        all_characters_in_dict = set()
        for speaker_targets in alias_dict.values():
            all_characters_in_dict.update(speaker_targets.keys())
        all_characters_in_dict.update(alias_dict.keys())  # 也包括所有speaker

        if target not in all_characters_in_dict:
            results.append({
                "used_name": used_name,
                "status": "unknown_target",
                "target": target,
                "reason": f"Target '{target}' not found in alias dictionary"
            })
            continue

        # 1. 常规规则检查
        correct_alias = speaker_alias_rules.get(target)
        if correct_alias and (used_name == correct_alias or used_name == target):
            results.append({
                "used_name": used_name,
                "status": "correct",
                "target": target
            })
            continue

        # 2. 特殊规则检查
        is_special_case = False
        if extra_info:
            for rule in extra_info:
                if (rule.get('speaker') == speaker_name and
                    rule.get('target') == target and
                        rule.get('alias') == used_name):
                    results.append({
                        "used_name": used_name,
                        "status": "special",
                        "target": target,
                        "conditions": rule.get('conditions', ['無特定条件'])
                    })
                    is_special_case = True
                    break

        if is_special_case:
            continue

        # 3. 错误情况
        results.append({
            "used_name": used_name,
            "status": "error",
            "target": target
        })

    return results


def _create_formatted_text_with_highlights(
    original_text: str,
    target_status_list: List[Dict[str, Any]]
) -> Any:
    """
    创建带有格式化高亮的文本对象

    Args:
        original_text: 原始文本
        target_status_list: target_used_name的状态列表

    Returns:
        openpyxl.styles.RichText对象
    """
    # 定义颜色
    GREEN = "00AA00"    # 正确 - 绿色
    RED = "FF0000"      # 错误 - 红色
    ORANGE = "FF8800"   # 特殊规则 - 橙色
    PURPLE = "8800AA"   # 未知speaker/target - 紫色

    # 创建富文本块列表
    text_blocks = []
    current_pos = 0

    # 按照used_name在文本中的出现位置排序
    name_positions = []
    for target_status in target_status_list:
        used_name = target_status["used_name"]
        pos = 0
        while True:
            pos = original_text.find(used_name, pos)
            if pos == -1:
                break
            name_positions.append({
                "start": pos,
                "end": pos + len(used_name),
                "status": target_status["status"],
                "used_name": used_name
            })
            pos += len(used_name)

    # 按位置排序
    name_positions.sort(key=lambda x: x["start"])

    # 创建普通字体（用于非格式化文本）
    normal_font = InlineFont()

    # 构建富文本块
    for name_pos in name_positions:
        # 添加前面的普通文本
        if current_pos < name_pos["start"]:
            normal_text = original_text[current_pos:name_pos["start"]]
            if normal_text:
                text_blocks.append(
                    TextBlock(text=normal_text, font=normal_font))

        # 添加格式化的target_used_name
        status = name_pos["status"]
        if status == "correct":
            color = GREEN
        elif status == "error":
            color = RED
        elif status == "special":
            color = ORANGE
        elif status in ["unknown_speaker", "unknown_target"]:
            color = PURPLE
        else:
            color = None

        if color:
            # 使用InlineFont而不是Font
            formatted_font = InlineFont(b=True, color=color)
            text_blocks.append(TextBlock(
                text=name_pos["used_name"],
                font=formatted_font
            ))
        else:
            # 普通称呼也需要字体
            text_blocks.append(
                TextBlock(text=name_pos["used_name"], font=normal_font))

        current_pos = name_pos["end"]

    # 添加剩余的普通文本
    if current_pos < len(original_text):
        remaining_text = original_text[current_pos:]
        if remaining_text:
            text_blocks.append(
                TextBlock(text=remaining_text, font=normal_font))

    # 创建富文本对象
    if text_blocks:
        return CellRichText(text_blocks)
    else:
        return original_text


async def _update_excel_with_findings(
    excel_file_path: str,
    text_parts: List[Dict[str, Any]],
    findings_results: List[AiReviewFindingEntry],
    s3_path: str,
    extra_info: List[Dict[str, Any]] = None,
    text_rpds: List[ReviewPointDefinitionVersionInDB] = None
) -> str:
    """
    将AI审查结果更新到Excel文件并上传到S3

    Args:
        excel_file_path: 本地Excel文件路径
        text_parts: 原始文本部分数据
        findings_results: AI审查结果
        s3_path: 原始S3路径
        extra_info: 特殊规则列表，用于Excel格式化时检测特殊规则

    Returns:
        str: 更新后的S3路径
    """
    try:
        # 从text_review RPD的reference_files中加载称呼表
        alias_dict = {}
        if text_rpds:
            alias_dict = _load_appellation_from_rpds(text_rpds)

        if not alias_dict:
            print("未能从RPD中加载称呼表，使用空字典")
            alias_dict = {}

        # 构建findings映射表，按(sheet_index, original_index)分组
        findings_map = {}
        for finding in findings_results:
            if finding.content_metadata:
                sheet_idx = finding.content_metadata.get('sheet_index', 0)
                orig_idx = finding.content_metadata.get('original_index', 0)
                key = (sheet_idx, orig_idx)

                if key not in findings_map:
                    findings_map[key] = []
                findings_map[key].append({
                    'description': finding.description,
                    'severity': finding.severity,
                    'tag': finding.tag,
                    'content_metadata': finding.content_metadata  # 保存完整的content_metadata
                })

        # 使用openpyxl加载Excel文件
        workbook = load_workbook(excel_file_path)

        # 与_download_and_process_text保持一致的列名
        text_col = "本文３４５６７８９０１２３４５６７８９０１２３４５６７８９０１２３４５６"

        # 按工作表处理
        sheet_names = workbook.sheetnames
        for sheet_index, sheet_name in enumerate(sheet_names):
            if sheet_name not in workbook.sheetnames:
                continue

            worksheet = workbook[sheet_name]

            # 读取该工作表的pandas DataFrame来获取列位置
            try:
                df = pd.read_excel(
                    excel_file_path, sheet_name=sheet_name, skiprows=11, engine='openpyxl')
                if text_col not in df.columns:
                    print(f"工作表 {sheet_name} 中未找到文本列，跳过")
                    continue

                text_col_idx = df.columns.get_loc(text_col)
                result_col_idx = text_col_idx + 2  # 在文本列右侧第2列写入结果

            except Exception as e:
                print(f"读取工作表 {sheet_name} 失败: {e}")
                continue

            # 遍历该工作表的所有文本部分
            for text_part in text_parts:
                if text_part.get('sheet_index') != sheet_index:
                    continue

                original_index = text_part.get('original_index', 0)
                key = (sheet_index, original_index)

                # 【新增功能】格式化原始文本列，标记target_used_name
                row_num = original_index + 13
                original_text = text_part.get('line', '')

                if key in findings_map:
                    findings = findings_map[key]

                    # 从findings中提取check_speaker结果
                    check_speaker_result = None
                    if key in findings_map and findings_map[key]:
                        # 获取content_metadata中的check_speaker结果
                        for finding in findings:
                            if finding.get('content_metadata'):
                                check_speaker_result = finding['content_metadata'].get(
                                    'check_speaker_result')
                                if check_speaker_result:
                                    break

                    if check_speaker_result and check_speaker_result.get('target_used_names'):
                        try:
                            # 分析target_used_names的状态
                            target_status_list = _analyze_target_used_names_status(
                                check_speaker_result,
                                alias_dict,
                                extra_info=extra_info or []  # 传入实际的特殊规则
                            )

                            # 创建格式化的文本
                            if target_status_list and original_text:
                                formatted_text = _create_formatted_text_with_highlights(
                                    original_text,
                                    target_status_list
                                )

                                # 更新文本列的内容
                                # pandas的列索引从0开始，openpyxl的列索引从1开始，所以需要+1
                                text_cell = worksheet.cell(
                                    row=row_num, column=text_col_idx + 1)

                                if hasattr(formatted_text, '__dict__'):
                                    print(
                                        f"  - formatted_text详情: {formatted_text.__dict__}")

                                text_cell.value = formatted_text

                                print(
                                    f"✅ 格式化文本完成 (sheet: {sheet_name}, row: {row_num}): 标记了 {len(target_status_list)} 个称呼")
                            else:
                                print(
                                    f"⚠️ 跳过格式化：target_status_list为空或original_text为空")
                        except Exception as format_error:
                            print(
                                f"❌ 格式化文本失败 (sheet: {sheet_name}, row: {row_num}): {format_error}")
                            import traceback
                            traceback.print_exc()
                            # 格式化失败时，保持原始文本不变
                    else:
                        print(f"⚠️ 跳过格式化：没有check_speaker结果或target_used_names为空")

                    # 构建检查结果消息
                    # check_messages = []
                    risk_severity_messages = []
                    alert_severity_messages = []
                    safe_severity_messages = []

                    for finding in findings:
                        severity = finding.get('severity', 'alert')
                        description = finding.get('description', '')
                        # tag = finding.get('tag', '')

                        if severity == 'risk':
                            risk_severity_messages.append(
                                f"【エラー】{description}")
                        elif severity == 'alert':
                            alert_severity_messages.append(
                                f"【警告】{description}")
                        elif severity == 'safe':
                            # 对于safe情况，只显示【OK】問題なし，不显示セリフ内容
                            safe_severity_messages.append("【OK】問題なし")
                        else:
                            # 向后兼容旧的severity值
                            if severity in ['high', 'medium', 'low']:
                                if severity == 'high':
                                    risk_severity_messages.append(
                                        f"【エラー】{description}")
                                elif severity in ['medium', 'low']:
                                    alert_severity_messages.append(
                                        f"【警告】{description}")
                            else:
                                # 未知severity默认为alert
                                alert_severity_messages.append(
                                    f"【警告】{description}")

                    # 组合消息：risk错误优先，然后是alert警告；若均无且存在safe，则输出safe
                    all_messages = risk_severity_messages + alert_severity_messages
                    if not all_messages and safe_severity_messages:
                        all_messages = safe_severity_messages
                    check_message = "\n".join(
                        all_messages) if all_messages else ""

                    if check_message or key in findings_map:
                        # 写入Excel单元格（行号需要+12因为跳过了11行）
                        try:
                            # +1列：写入detected_speaker
                            speaker_info = ""
                            if check_speaker_result:
                                detected_speaker = check_speaker_result.get(
                                    'detected_speaker', '')
                                if detected_speaker:
                                    speaker_info = detected_speaker

                            # +2列：写入target_list
                            target_info = []
                            if check_speaker_result:
                                target_list = check_speaker_result.get(
                                    'target_list', [])
                                if target_list:
                                    target_info = target_list

                            # 写入数据到不同列
                            if speaker_info:
                                worksheet.cell(
                                    row=row_num, column=result_col_idx + 1, value=speaker_info)
                            if target_info:
                                # target_info是list，转换为字符串写入Excel
                                target_str = ", ".join(target_info) if isinstance(
                                    target_info, list) else str(target_info)
                                worksheet.cell(
                                    row=row_num, column=result_col_idx + 2, value=target_str)
                            if check_message:
                                worksheet.cell(
                                    row=row_num, column=result_col_idx + 3, value=check_message)

                            # print(f"更新工作表 {sheet_name}, 行 {row_num}: speaker={speaker_info}, targets={target_info}, message={check_message[:30]}...")
                        except Exception as e:
                            print(
                                f"写入Excel单元格失败 (sheet: {sheet_name}, row: {row_num}): {e}")

        # 保存Excel文件
        # print(f"📁 开始保存Excel文件: {excel_file_path}")
        workbook.save(excel_file_path)
        print(f"✅ Excel文件已更新并保存: {excel_file_path}")

        try:
            # 重新加载验证
            verify_workbook = load_workbook(excel_file_path)
            print(
                f"📋 验证：重新加载Excel文件成功，工作表数量: {len(verify_workbook.sheetnames)}")
            verify_workbook.close()
        except Exception as verify_error:
            print(f"⚠️ 验证Excel文件时出错: {verify_error}")

        # 上传到S3
        new_s3_path = await _upload_updated_excel_to_s3(excel_file_path, s3_path)
        return new_s3_path

    except Exception as e:
        print(f"更新Excel文件失败: {e}")
        raise e


async def _upload_updated_excel_to_s3(local_file_path: str, original_s3_path: str) -> str:
    """
    将更新后的Excel文件上传到S3，直接替换原文件

    Args:
        local_file_path: 本地文件路径
        original_s3_path: 原始S3路径（支持完整S3 URL或相对路径）

    Returns:
        str: S3路径（与原文件路径相同）
    """
    try:
        from aimage_supervision.settings import AWS_BUCKET_NAME

        # 处理S3路径：支持完整URL和相对路径
        if original_s3_path.startswith("s3://"):
            # 完整S3 URL格式：s3://bucket/path/file.xlsx
            s3_parts = original_s3_path.replace("s3://", "").split("/", 1)
            bucket_name = s3_parts[0]
            object_key = s3_parts[1] if len(s3_parts) > 1 else ""
        else:
            # 相对路径格式：path/file.xlsx
            bucket_name = AWS_BUCKET_NAME
            object_key = original_s3_path

        # 直接使用原文件名（替换策略）
        new_object_key = object_key  # 直接使用原文件的object_key，实现文件替换

        # 上传到S3（使用与其他函数相同的配置）
        from aimage_supervision.settings import (AWS_ACCESS_KEY_ID, AWS_REGION,
                                                 AWS_SECRET_ACCESS_KEY)

        s3_client = boto3.client(
            's3',
            region_name=AWS_REGION,
            aws_access_key_id=AWS_ACCESS_KEY_ID,
            aws_secret_access_key=AWS_SECRET_ACCESS_KEY
        )
        # 检查原文件是否存在
        try:
            s3_client.head_object(Bucket=bucket_name, Key=new_object_key)
            print(f"🔄 将直接替换原文件: {new_object_key}")
        except s3_client.exceptions.NoSuchKey:
            print(f"⚠️ 原文件不存在，将创建新文件: {new_object_key}")
        except Exception as e:
            print(f"检查原文件存在性时出错: {e}")

        # 上传文件（直接替换原文件）
        s3_client.upload_file(local_file_path, bucket_name, new_object_key)

        new_s3_path = f"s3://{bucket_name}/{new_object_key}"
        print(f"✅ Excel文件已成功替换原文件到S3: {new_s3_path}")

        return new_s3_path

    except Exception as e:
        print(f"上传Excel文件到S3失败: {e}")
        raise e


async def _build_final_ai_review_data(
    ai_review_orm: AiReview,
    finding_entries: List[AiReviewFindingEntry]
) -> AiReviewInDB:
    """构建最终的AI审查数据并更新数据库"""
    finding_entry_schemas = [
        AiReviewFindingEntrySchema.from_orm(fe) for fe in finding_entries]

    # 构建检测元素的摘要数据
    detected_elements_summary: Dict[str, Any] = {
        "description": f"AI审查发现了 {len(finding_entries)} 个问题",
        "elements": []
    }

    # 根据findings生成检测元素摘要
    for finding_entry in finding_entries:
        # 使用tag作为name，如果没有tag则使用截断的description
        element_name = finding_entry.tag if finding_entry.tag else finding_entry.description[
            :50]

        # 确保符合AiDetectedElement schema的必需字段
        element = {
            "name": element_name,
            "confidence": 0.8,
            "label": "object",  # 必需字段：使用'object'作为默认label
            # 必需字段：使用finding的area或默认值
            "area": finding_entry.area or {"x": 0, "y": 0, "width": 0, "height": 0},
            "character_id": None
        }
        detected_elements_summary["elements"].append(element)

    # 更新数据库中的ai_review_orm对象
    ai_review_orm.ai_review_output_json = detected_elements_summary
    ai_review_orm.review_timestamp = ai_review_orm.created_at
    await ai_review_orm.save()

    final_review_data = AiReviewInDB(
        id=ai_review_orm.id,
        created_at=ai_review_orm.created_at,
        updated_at=ai_review_orm.updated_at,
        subtask_id=ai_review_orm.subtask.id,
        version=ai_review_orm.version,
        is_latest=ai_review_orm.is_latest,
        review_timestamp=ai_review_orm.review_timestamp,
        initiated_by_user_id=ai_review_orm.initiated_by_user.id if ai_review_orm.initiated_by_user else None,
        last_modified_by_user_id=ai_review_orm.last_modified_by_user.id if ai_review_orm.last_modified_by_user else None,
        findings=finding_entry_schemas
    )

    return final_review_data


async def get_subtasks_for_tasks(task_ids: List[UUID]) -> List[Subtask]:
    """
    Retrieves all subtask objects associated with a list of task IDs.
    """
    subtasks = await Subtask.filter(task_id__in=task_ids).all()
    return subtasks


async def initiate_ai_review_for_subtask(
    subtask_id: UUID,
    initiated_by_user_id: UUID,
    cr_check: Optional[bool] = False,
    rpd_ids: Optional[List[UUID]] = None,
    mode: AiReviewMode = AiReviewMode.QUALITY
) -> AiReviewInDB:
    """
    为指定的子任务启动完整的AI审查流程。
    这包括:
    1. 获取子任务数据和图像
    2. 对图像执行元素检测
    3. 创建新的AiReview记录（版本化）
    4. 为每个活跃的ReviewPointDefinitionVersion生成和存储findings
    5. 返回综合的AiReview数据
    """
    ai_review_orm = None
    try:
        # 1. 创建AI Review记录 - 状态：PENDING
        ai_review_orm = await _create_new_ai_review_version(subtask_id, initiated_by_user_id)

        # 2. 更新状态为PROCESSING
        await _update_processing_status(
            ai_review_orm,
            AiReviewProcessingStatus.PROCESSING
        )

        # 3. 执行AI处理（基于内容类型的分支处理）
        subtask = await _validate_and_fetch_subtask(subtask_id)

        # 检查subtask的内容类型
        content_type = subtask.content.task_type if subtask.content else SubtaskType.PICTURE

        active_rpd_versions = await get_active_review_point_versions(
            project_id=subtask.task.project.id,
            cr_check=cr_check,
            rpd_ids=rpd_ids
        )

        if not active_rpd_versions:
            print(
                f"Warning: No active/specified Review Point Definition Versions found for project {subtask.task.project.id}. "
                f"AI Review for subtask {subtask_id} will have no findings from RPDs."
            )

        if content_type == SubtaskType.PICTURE:
            # 图片处理逻辑（原有逻辑）
            # 检查中断信号
            if await _check_cancellation_signal(ai_review_orm):
                await _update_processing_status(ai_review_orm, AiReviewProcessingStatus.CANCELLED)
                return await _build_final_ai_review_data(ai_review_orm, [])

            image_bytes, image_width, image_height = await _download_and_process_image(
                subtask_id, subtask.content.s3_path
            )
            content_metadata: Dict[str, Any] = {}
            all_finding_entries = await _generate_findings_for_all_rpd_versions(
                image_bytes, image_width, image_height, content_type, content_metadata, active_rpd_versions, ai_review_orm, mode,
            )

        elif content_type == SubtaskType.VIDEO:
            # 视频处理逻辑 - 并行处理所有视频帧
            # 检查中断信号
            if await _check_cancellation_signal(ai_review_orm):
                await _update_processing_status(ai_review_orm, AiReviewProcessingStatus.CANCELLED)
                return await _build_final_ai_review_data(ai_review_orm, [])

            print(f"Processing video content for subtask {subtask_id}")
            video_data, video_width, video_height = await _download_and_process_video(subtask_id, subtask.content.s3_path)

            # 使用并行处理函数处理所有视频帧
            all_finding_entries = await _generate_findings_for_video_frames_parallel(
                video_data, video_width, video_height, content_type,
                active_rpd_versions, ai_review_orm, mode
            )

        elif content_type in [SubtaskType.TEXT, SubtaskType.WORD, SubtaskType.EXCEL]:
            # 文本/文档处理逻辑 - 并行处理所有文本部分
            # 检查中断信号
            if await _check_cancellation_signal(ai_review_orm):
                await _update_processing_status(ai_review_orm, AiReviewProcessingStatus.CANCELLED)
                return await _build_final_ai_review_data(ai_review_orm, [])

            print(
                f"Processing {content_type.value} content for subtask {subtask_id}")

            try:
                text_parts, local_excel_path = await _download_and_process_text(subtask_id, subtask.content.s3_path)
                if text_parts:
                    # 使用并行处理函数处理所有文本部分，并更新Excel文件
                    # 从text_review RPD的special_rules字段获取特殊规则
                    extra_info = []
                    text_rpds = [
                        v for v in active_rpd_versions if v.parent_key == 'text_review']
                    for rpd in text_rpds:
                        if rpd.special_rules and isinstance(rpd.special_rules, list):
                            # special_rules现在是JSON格式的列表
                            for rule in rpd.special_rules:
                                if isinstance(rule, dict) and all(key in rule for key in ['speaker', 'target', 'alias']):
                                    extra_info.append({
                                        "speaker": rule.get('speaker', ''),
                                        "target": rule.get('target', ''),
                                        "alias": rule.get('alias', ''),
                                        "conditions": rule.get('conditions', []),
                                        "rpd_title": rpd.title
                                    })
                    print(f"从RPD获取到 {len(extra_info)} 条特殊规则")

                    all_finding_entries = await _generate_findings_for_text_parts_parallel(
                        text_parts, content_type, active_rpd_versions, ai_review_orm, mode,
                        local_excel_path, subtask.content.s3_path, extra_info=extra_info)
                else:
                    print(f"Error: {text_parts}")
                    all_finding_entries = []
                    # 清理临时文件
                    if 'local_excel_path' in locals() and os.path.exists(local_excel_path):
                        os.unlink(local_excel_path)

            except NotImplementedError:
                print(f"Error: {NotImplementedError}")
                all_finding_entries = []

        elif content_type == SubtaskType.AUDIO:
            # 音频处理逻辑 - 并行处理所有音频片段
            # 检查中断信号
            if await _check_cancellation_signal(ai_review_orm):
                await _update_processing_status(ai_review_orm, AiReviewProcessingStatus.CANCELLED)
                return await _build_final_ai_review_data(ai_review_orm, [])

            print(f"Processing audio content for subtask {subtask_id}")

            try:
                # TODO: 实现音频下载和分割逻辑
                # audio_data, audio_metadata = await _download_and_process_audio(subtask_id, subtask.content.s3_path)
                # audio_segments = await _split_audio_into_segments(audio_data, audio_metadata)

                # 暂时创建一个示例音频片段列表，展示并行处理的用法
                # 实际实现时，audio_segments应该来自于音频分割函数
                audio_segments = []  # 空列表，因为音频处理功能尚未实现

                if audio_segments:
                    ...
                    # 使用并行处理函数处理所有音频片段
                    # all_finding_entries = await _generate_findings_for_audio_segments_parallel(
                    #     audio_segments, content_type, active_rpd_versions, ai_review_orm, mode, max_concurrent=5
                    # )
                else:
                    print(f"音频处理功能尚未完整实现，跳过音频内容")
                    all_finding_entries = []  # 暂时返回空列表，直到实现音频处理逻辑

            except NotImplementedError:
                print(f"音频处理功能尚未实现，跳过音频内容")
                all_finding_entries = []

        else:
            # 未知类型处理
            print(
                f"Unknown content type {content_type} for subtask {subtask_id}, skipping AI review")
            all_finding_entries = []

        # 4. 构建最终数据
        final_data = await _build_final_ai_review_data(ai_review_orm, all_finding_entries)

        # 5. 更新状态为COMPLETED
        await _update_processing_status(
            ai_review_orm,
            AiReviewProcessingStatus.COMPLETED
        )

        return final_data

    except Exception as e:
        # 6. 异常处理：更新状态为FAILED并回退到上一个版本
        if ai_review_orm:
            await _update_processing_status(
                ai_review_orm,
                AiReviewProcessingStatus.FAILED,
                error_message=str(e)
            )

            # 执行回退逻辑
            logger.info(
                f"AI review failed for subtask {subtask_id}, attempting rollback...")
            # 确保subtask关系已加载
            if not hasattr(ai_review_orm, 'subtask') or ai_review_orm.subtask is None:
                await ai_review_orm.fetch_related('subtask')
            await _rollback_failed_review(ai_review_orm)

        raise


# --- Service functions for retrieving AI Review data (Phase 3) ---

async def _map_ai_review_orm_to_schema(review_orm: AiReview) -> AiReviewSchema:
    """
    Helper function to map an AiReview ORM model to the AiReviewSchema Pydantic model,
    populating findings and detected_elements.
    """
    # Ensure findings and related objects are loaded
    await review_orm.fetch_related('findings', 'subtask', 'initiated_by_user', 'last_modified_by_user')

    # finding_schemas will be List[AiReviewFindingEntryInDB]
    finding_schemas_in_db = [AiReviewFindingEntryInDB.from_orm(
        f) for f in review_orm.findings]

    # Debug: Check what's in the ORM objects and schema objects
    # for f in review_orm.findings:
    #     print(f"DEBUG ORM Finding {f.id}: is_fixed = {f.is_fixed}")

    # for f_in_db in finding_schemas_in_db:
    #     print(
    #         f"DEBUG Schema Finding {f_in_db.id}: is_fixed = {f_in_db.is_fixed}")
    #     print(f"DEBUG model_dump: {f_in_db.model_dump()}")

    # Convert to List[AiReviewFindingEntry] as expected by AiReviewSchema

    final_findings_list: List[AiReviewFindingEntrySchema] = [
        AiReviewFindingEntrySchema(**f_in_db.model_dump()) for f_in_db in finding_schemas_in_db
    ]

    # Debug: Check final findings list
    # for final_finding in final_findings_list:
    #     print(
    #         f"DEBUG Final Finding {final_finding.id}: is_fixed = {final_finding.is_fixed}")
    #     print(f"DEBUG Final Finding model_dump: {final_finding.model_dump()}")

    parsed_summary_object: Optional[AiDetectedElements] = None
    if review_orm.ai_review_output_json and isinstance(review_orm.ai_review_output_json, dict):
        try:
            parsed_summary_object = AiDetectedElements.model_validate(
                review_orm.ai_review_output_json)
        except Exception as e:
            print(
                f"Warning: Could not parse ai_review_output_json for AiReview ID {review_orm.id}: {e}")
            # parsed_summary_object remains None if validation fails

    final_detected_elements_list: Optional[List[AiDetectedElement]] = None
    if parsed_summary_object:
        final_detected_elements_list = parsed_summary_object.elements

    subtask_id_val = review_orm.subtask.id
    initiated_by_user_id_val = review_orm.initiated_by_user.id if review_orm.initiated_by_user else None
    last_modified_by_user_id_val = review_orm.last_modified_by_user.id if review_orm.last_modified_by_user else None

    return AiReviewSchema(
        id=review_orm.id,
        subtask_id=subtask_id_val,
        version=review_orm.version,
        is_latest=review_orm.is_latest,
        review_timestamp=review_orm.review_timestamp,
        initiated_by_user_id=initiated_by_user_id_val,
        last_modified_by_user_id=last_modified_by_user_id_val,
        created_at=review_orm.created_at,
        updated_at=review_orm.updated_at,
        findings=final_findings_list,
        detected_elements=final_detected_elements_list,
        detected_elements_summary=parsed_summary_object,
        # 新增状态字段
        processing_status=review_orm.processing_status.value if review_orm.processing_status else None,
        error_message=review_orm.error_message,
        processing_started_at=review_orm.processing_started_at,
        processing_completed_at=review_orm.processing_completed_at
    )


async def get_ai_review_by_id(ai_review_id: UUID) -> Optional[AiReviewSchema]:
    """
    Retrieves a specific AI Review by its ID, including its findings and parsed detected elements.
    """
    try:
        review_orm = await AiReview.get_or_none(id=ai_review_id)
        if not review_orm:
            return None
        return await _map_ai_review_orm_to_schema(review_orm)
    except DoesNotExist:  # Should be caught by get_or_none, but as a safeguard
        return None


async def get_latest_ai_review_for_subtask(subtask_id: UUID) -> Optional[AiReviewSchema]:
    """
    Retrieves the latest AI Review for a given subtask, including findings and parsed detected elements.
    """
    review_orm = await AiReview.filter(subtask_id=subtask_id, is_latest=True).first()
    if not review_orm:
        return None
    return await _map_ai_review_orm_to_schema(review_orm)


async def list_ai_reviews_for_subtask(subtask_id: UUID) -> List[AiReviewSchema]:
    """
    List all AI reviews for a subtask, ordered by version descending.
    """
    review_orms = await AiReview.filter(subtask_id=subtask_id).order_by('-version').prefetch_related('findings')
    review_schemas = []
    for review_orm in review_orms:
        review_schemas.append(await _map_ai_review_orm_to_schema(review_orm))
    return review_schemas


class CharacterPredictionResponse(BaseModel):
    character_count: int
    character_name: List[str]


async def predict_character_for_subtask(
    subtask: Subtask, character_candidates: List[Character],
) -> Tuple[List[Character] | None, float]:
    """预测子任务中的角色

    Args:
        subtask: 子任务对象
        character_candidates: 候选角色列表
    """

    # 为这个任务创建独立的 Gemini 客户端
    task_gemini_client = genai.Client(api_key=os.getenv(
        'GEMINI_API_KEY')) if os.getenv('GEMINI_API_KEY') else None
    if not task_gemini_client:
        raise ValueError("GEMINI_API_KEY not set. Cannot predict character.")

    # 检查角色候选列表中是否有有效的图片路径
    valid_characters = []
    for character in character_candidates:
        if not character.image_path or character.image_path.strip() == "":
            print(
                f"Warning: Character {character.name} (ID: {character.id}) has no image_path, skipping...")
            continue
        valid_characters.append(character)

    if not valid_characters:
        raise ValueError(
            "No characters with valid image paths found in the project")

    character_image_prompts = []
    for character in valid_characters:
        try:
            image_bytes = download_file_content_from_s3_sync(
                character.image_path)
            character_image_prompts.extend([
                f"Character: {character.name}",
                types.Part.from_bytes(data=image_bytes, mime_type='image/jpeg')
            ])
        except Exception as e:
            print(
                f"Error downloading image for character {character.name} (path: {character.image_path}): {e}")
            continue

    if not character_image_prompts:
        raise ValueError("Failed to download images for any characters")

    # get subtask image from s3
    subtask_image_path = subtask.content.s3_path
    if not subtask_image_path or subtask_image_path.strip() == "":
        raise ValueError("Subtask has no valid s3_path for content")

    subtask_image_bytes = download_file_content_from_s3_sync(
        subtask_image_path)

    # 使用新的工具函数自动处理图片或视频，如果是视频会自动提取第一帧
    subtask_image_bytes = process_image_or_video_bytes(subtask_image_bytes)

    model_name = "gemini-2.5-flash"
    contents = ["The following is a list of characters and their images: "] + \
        character_image_prompts + \
        ["Please predict the characters in the following image",
         types.Part.from_bytes(
             data=subtask_image_bytes, mime_type='image/jpeg'),
         "Return the number of characters in the image and the names of the characters in a list.",
         "The subtask content is: " + str(subtask.content)]
    response = task_gemini_client.models.generate_content(
        model=model_name,
        contents=contents,
        config=genai.types.GenerateContentConfig(
            response_mime_type='application/json',
            response_schema=CharacterPredictionResponse,
            system_instruction="You are a character analysis AI. Recognize how many characters are in the image and return the names of the characters.",
            temperature=0.0,
        ),
    )

    character_count = response.parsed.character_count
    character_names = response.parsed.character_name
    print(
        f"Character count: {character_count}, Character names: {character_names}")

    if character_count != len(character_names):
        raise ValueError(
            f"Character count {character_count} does not match the number of character names {len(character_names)}")

    predicted_characters = []
    for character_name in character_names:
        for character in valid_characters:
            # print(f"Character name: {character_name}")
            character_name = character_name.strip()
            if character.name.lower() == character_name.lower():
                predicted_characters.append(character)
    print(f"Predicted characters: {predicted_characters}")
    if len(predicted_characters) == 0:
        return None, 1.0

    return predicted_characters, 1.0


async def _validate_and_fetch_subtask(subtask_id: UUID) -> Subtask:
    """验证并获取subtask数据"""
    try:
        subtask = await Subtask.get(id=subtask_id).prefetch_related('task__project')
    except DoesNotExist:
        raise ValueError(f"Subtask with ID {subtask_id} not found.")

    if not subtask.content or not subtask.content.s3_path:
        raise ValueError(
            f"Subtask {subtask_id} does not have a valid s3_path in its content.")

    return subtask


async def _download_and_process_image(subtask_id: UUID, s3_path: str) -> tuple[bytes, int, int]:
    """下载并处理图像，返回图像字节数据和尺寸"""
    image_path_local = None

    try:
        image_path_local = await download_image_from_s3(s3_path)
        with open(image_path_local, 'rb') as f:
            image_bytes = f.read()

        image = Image.open(image_path_local)
        image_width, image_height = image.size

        return image_bytes, image_width, image_height

    except Exception as e:
        # Clean up if download or read failed
        if image_path_local and os.path.exists(image_path_local):
            os.remove(image_path_local)
        raise ValueError(
            f"Failed to download or read image for subtask {subtask_id} from {s3_path}: {e}")


async def _download_and_process_video(subtask_id: UUID, s3_path: str) -> Tuple[List[Dict[str, Any]], int, int]:
    """下载并处理视频，返回视频字节数据和尺寸"""
    video_path_local = None

    try:
        video_path_local = await download_video_from_s3(s3_path)
        video_frames_info = detect_video_scenes_and_extract_frames(
            video_path_local)
        video_width, video_height = video_frames_info["width"], video_frames_info["height"]
        return video_frames_info["scenes"], video_width, video_height

    except Exception as e:
        # Clean up if download or read failed
        if video_path_local and os.path.exists(video_path_local):
            os.remove(video_path_local)
        raise ValueError(
            f"Failed to download or read video for subtask {subtask_id} from {s3_path}: {e}")


async def _check_cancellation_signal(ai_review_orm: AiReview) -> bool:
    """检查是否有中断信号"""
    await ai_review_orm.refresh_from_db(fields=['should_cancel'])
    if ai_review_orm.should_cancel:
        logger.info(
            f"Cancellation signal detected for AI review {ai_review_orm.id}")
    return ai_review_orm.should_cancel


async def _rollback_failed_review(failed_review_orm: AiReview) -> None:
    """
    回退失败的AI Review到上一个版本

    Args:
        failed_review_orm: 失败的AI Review ORM对象
    """
    try:
        # 确保subtask关系已加载
        await failed_review_orm.fetch_related('subtask')
        subtask_id = failed_review_orm.subtask.id
        failed_version = failed_review_orm.version

        logger.info(
            f"Rolling back failed AI review version {failed_version} for subtask {subtask_id}")

        # 查找上一个成功的版本
        logger.info(
            f"Searching for successful versions before version {failed_version}")
        all_previous_reviews = await AiReview.filter(
            subtask=failed_review_orm.subtask,
            version__lt=failed_version
        ).order_by('-version')

        logger.info(f"Found {len(all_previous_reviews)} previous versions:")
        for review in all_previous_reviews:
            logger.info(
                f"  Version {review.version}: status={review.processing_status}, is_latest={review.is_latest}")

        previous_successful_review = await AiReview.filter(
            subtask=failed_review_orm.subtask,
            version__lt=failed_version,
            processing_status=AiReviewProcessingStatus.COMPLETED.value
        ).order_by('-version').first()

        if previous_successful_review:
            # 有上一个成功版本，回退到该版本
            logger.info(
                f"Found previous successful version {previous_successful_review.version}, rolling back to it")

            # 将失败版本的is_latest设置为False
            failed_review_orm.is_latest = False
            await failed_review_orm.save(update_fields=['is_latest'])

            # 将上一个成功版本设置为latest
            previous_successful_review.is_latest = True
            await previous_successful_review.save(update_fields=['is_latest'])

            logger.info(
                f"Successfully rolled back to version {previous_successful_review.version}")

        else:
            # 没有上一个成功版本，回退到"无监修状态"
            logger.info(
                "No previous successful version found, rolling back to no-review state")

            # 将失败版本的is_latest设置为False，这样get_latest_ai_review_for_subtask会返回None
            failed_review_orm.is_latest = False
            await failed_review_orm.save(update_fields=['is_latest'])

            logger.info("Successfully rolled back to no-review state")

    except Exception as rollback_error:
        logger.error(
            f"Failed to rollback AI review {failed_review_orm.id}: {rollback_error}")
        import traceback
        logger.error(f"Rollback traceback: {traceback.format_exc()}")
        # 回退失败不应该影响主要的错误处理流程，但我们需要记录详细错误
        pass


async def _wait_for_tasks_with_cancellation(
    tasks: List[asyncio.Task],
    ai_review_orm: AiReview,
    check_interval: float = 2.0
) -> List[asyncio.Task]:
    """
    可中断的任务等待机制
    每隔check_interval秒检查一次中断信号，如果检测到中断则取消所有任务
    返回已完成的任务列表
    """
    done = set()
    pending = set(tasks)

    while pending:
        # 检查中断信号
        if await _check_cancellation_signal(ai_review_orm):
            print("检测到中断信号，取消所有待处理任务")
            # 取消所有待处理的任务
            for task in pending:
                task.cancel()
            # 等待已取消的任务完成清理
            await asyncio.gather(*pending, return_exceptions=True)
            print("所有任务已取消")
            return []

        # 等待至少一个任务完成，或者超时检查中断信号
        completed, pending = await asyncio.wait(
            pending,
            return_when=asyncio.FIRST_COMPLETED,
            timeout=check_interval
        )

        done.update(completed)

        if completed:
            print(f"已完成 {len(done)} / {len(tasks)} 个任务")

    return list(done)


async def _update_processing_status(
    ai_review_orm: AiReview,
    status: AiReviewProcessingStatus,
    error_message: Optional[str] = None
) -> None:
    """更新AI Review处理状态的统一函数"""
    ai_review_orm.processing_status = status

    if status == AiReviewProcessingStatus.PROCESSING:
        ai_review_orm.processing_started_at = datetime.now(timezone.utc)
    elif status in [AiReviewProcessingStatus.COMPLETED, AiReviewProcessingStatus.FAILED, AiReviewProcessingStatus.CANCELLED]:
        ai_review_orm.processing_completed_at = datetime.now(timezone.utc)
        if error_message:
            ai_review_orm.error_message = error_message

    await ai_review_orm.save()


async def _create_new_ai_review_version(subtask_id: UUID, initiated_by_user_id: UUID) -> AiReview:
    """创建新的AI审查版本，状态默认为PENDING"""
    # 获取subtask和user对象
    subtask = await Subtask.get(id=subtask_id)
    user = await User.get(id=initiated_by_user_id)

    previous_reviews = await AiReview.filter(subtask=subtask).order_by("-version")

    # 版本管理策略：保留历史版本，只更新is_latest标记
    if previous_reviews:
        print(f"找到 {len(previous_reviews)} 个旧的AI Review记录，将更新is_latest标记...")
        # 将所有旧版本的is_latest设置为False
        await AiReview.filter(subtask=subtask).update(is_latest=False)

        # 计算新版本号
        latest_version = previous_reviews[0].version
        new_review_version = latest_version + 1
        print(f"创建新版本: {new_review_version}")
    else:
        # 首次创建，版本号为1
        new_review_version = 1
        print("首次创建AI Review，版本号为1")

    ai_review_orm = await AiReview.create(
        subtask=subtask,
        version=new_review_version,
        is_latest=True,
        ai_review_output_json={},
        initiated_by_user=user,
        processing_status=AiReviewProcessingStatus.PENDING  # 明确设置初始状态
    )

    return ai_review_orm


async def get_relevant_findings_for_subtask(subtask_id: UUID) -> List[AiReviewFindingEntry]:
    """
    获取子任务的相关发现：包括最新版本的所有发现 + 历史版本中被标记为 is_fixed 的发现
    这种方法避免了遍历所有历史版本，只通过一次查询获取所有相关数据
    """
    # 使用单个查询获取所有相关的发现条目
    # 条件：属于该 subtask 且 (是最新版本 OR 被标记为 is_fixed)
    findings = await AiReviewFindingEntry.filter(
        ai_review__subtask_id=subtask_id
    ).filter(
        Q(ai_review__is_latest=True) | Q(is_fixed=True)
    ).prefetch_related(
        'ai_review',
        'review_point_definition_version',
        'review_point_definition_version__review_point_definition'
    ).order_by('-ai_review__version', 'created_at')

    return findings


async def get_findings_summary_for_subtask(subtask_id: UUID) -> Optional[AiReviewSchema]:
    """
    获取子任务的发现摘要，返回与 get_latest_ai_review_for_subtask 相同的格式，
    但 findings 字段包含优化过的数据：最新版本的findings + 历史版本中标记为 is_fixed 的findings
    这避免了扫描所有历史版本，提供更好的性能。
    """
    # 获取最新的 AI review ORM 对象
    latest_review_orm = await AiReview.filter(subtask_id=subtask_id, is_latest=True).first()
    if not latest_review_orm:
        return None

    # 获取相关的 findings（最新版本 + 历史版本中 is_fixed 的）
    relevant_findings = await get_relevant_findings_for_subtask(subtask_id)

    # 确保相关对象已加载
    await latest_review_orm.fetch_related('subtask', 'initiated_by_user', 'last_modified_by_user')

    # 将相关 findings 转换为 schema 格式
    finding_schemas_in_db = [
        AiReviewFindingEntryInDB.from_orm(f) for f in relevant_findings]

    final_findings_list: List[AiReviewFindingEntrySchema] = [
        AiReviewFindingEntrySchema(**f_in_db.model_dump()) for f_in_db in finding_schemas_in_db
    ]

    # 解析 detected_elements_summary
    parsed_summary_object: Optional[AiDetectedElements] = None
    if latest_review_orm.ai_review_output_json and isinstance(latest_review_orm.ai_review_output_json, dict):
        try:
            parsed_summary_object = AiDetectedElements.model_validate(
                latest_review_orm.ai_review_output_json)
        except Exception as e:
            print(
                f"Warning: Could not parse ai_review_output_json for AiReview ID {latest_review_orm.id}: {e}")

    final_detected_elements_list: Optional[List[AiDetectedElement]] = None
    if parsed_summary_object:
        final_detected_elements_list = parsed_summary_object.elements

    subtask_id_val = latest_review_orm.subtask.id
    initiated_by_user_id_val = latest_review_orm.initiated_by_user.id if latest_review_orm.initiated_by_user else None
    last_modified_by_user_id_val = latest_review_orm.last_modified_by_user.id if latest_review_orm.last_modified_by_user else None

    return AiReviewSchema(
        id=latest_review_orm.id,
        subtask_id=subtask_id_val,
        version=latest_review_orm.version,
        is_latest=latest_review_orm.is_latest,
        review_timestamp=latest_review_orm.review_timestamp,
        initiated_by_user_id=initiated_by_user_id_val,
        last_modified_by_user_id=last_modified_by_user_id_val,
        created_at=latest_review_orm.created_at,
        updated_at=latest_review_orm.updated_at,
        findings=final_findings_list,  # 使用优化过的 findings
        detected_elements=final_detected_elements_list,
        detected_elements_summary=parsed_summary_object,
        # 新增状态字段
        processing_status=latest_review_orm.processing_status.value if latest_review_orm.processing_status else None,
        error_message=latest_review_orm.error_message,
        processing_started_at=latest_review_orm.processing_started_at,
        processing_completed_at=latest_review_orm.processing_completed_at
    )


async def update_finding_fixed_status(finding_id: UUID, is_fixed: bool) -> AiReviewFindingEntry:
    """
    更新发现条目的 is_fixed 状态
    """
    finding = await AiReviewFindingEntry.get_or_none(id=finding_id)
    if not finding:
        raise ValueError(f"Finding with ID {finding_id} not found")

    finding.is_fixed = is_fixed
    await finding.save()

    return finding


async def update_finding_content(
    finding_id: UUID,
    description: Optional[str] = None,
    severity: Optional[str] = None,
    suggestion: Optional[str] = None
) -> AiReviewFindingEntry:
    """
    更新AI审查发现条目的内容

    Args:
        finding_id: 发现条目的ID
        description: 新的描述（可选）
        severity: 新的严重程度（可选）
        suggestion: 新的建议（可选）

    Returns:
        AiReviewFindingEntry: 更新后的发现条目

    Raises:
        ValueError: 如果找不到发现条目
    """
    finding = await AiReviewFindingEntry.get_or_none(id=finding_id)
    if not finding:
        raise ValueError("Finding not found")

    # Update only provided fields
    if description is not None:
        finding.description = description
    if severity is not None:
        finding.severity = severity
    if suggestion is not None:
        finding.suggestion = suggestion

    # Update the updated_at timestamp
    finding.updated_at = datetime.now(timezone.utc)

    await finding.save()
    return finding


async def update_finding_bounding_box(
    finding_id: UUID,
    area: dict
) -> AiReviewFindingEntry:
    """
    更新AI审查发现条目的边界框

    Args:
        finding_id: 发现条目的ID
        area: 新的边界框数据（包含x, y, width, height）

    Returns:
        AiReviewFindingEntry: 更新后的发现条目

    Raises:
        ValueError: 如果找不到发现条目
    """
    finding = await AiReviewFindingEntry.get_or_none(id=finding_id)
    if not finding:
        raise ValueError("Finding not found")

    # Update the area field
    finding.area = area

    # Update the updated_at timestamp
    finding.updated_at = datetime.now(timezone.utc)

    await finding.save()
    return finding


# --- Direct Image Analysis for Testing (New Function) ---

async def analyze_image_with_rpds(
    image_bytes: bytes,
    image_filename: str,
    rpd_ids: Optional[List[UUID]] = None,
    cr_check: bool = False,
    mode: AiReviewMode = AiReviewMode.QUALITY,
    user_id: UUID = None,
    project_id: Optional[UUID] = None
) -> Dict[str, Any]:
    """
    直接分析图片与指定的RPDs，不创建数据库记录

    这是专为RPD测试和开发设计的方法，特点：
    - 直接处理图片字节数据
    - 不创建AI Review或subtask记录
    - 快速返回分析结果
    - 适合测试和原型开发

    Args:
        image_bytes: 图片的字节数据
        image_filename: 图片文件名（用于日志）
        rpd_ids: 要使用的RPD ID列表
        cr_check: 是否启用版权检查
        mode: 分析模式（quality/speed）
        user_id: 用户ID（用于日志）
        project_id: 项目ID（用于筛选RPD）

    Returns:
        Dict包含分析结果：
        {
            "success": True/False,
            "findings": [...],
            "rpd_count": int,
            "processing_time": float,
            "image_info": {...}
        }
    """
    start_time = time.time()

    try:
        # 1. 处理和验证图片
        print(f"开始分析图片: {image_filename}")

        # 获取图片尺寸
        image_width, image_height = await _get_image_dimensions(image_bytes)

        # 2. 获取活跃的RPD版本
        active_rpd_versions = await get_active_review_point_versions(
            project_id=project_id,
            cr_check=cr_check,
            rpd_ids=rpd_ids
        )

        if not active_rpd_versions:
            return {
                "success": True,
                "findings": [],
                "rpd_count": 0,
                "processing_time": time.time() - start_time,
                "image_info": {
                    "filename": image_filename,
                    "width": image_width,
                    "height": image_height
                },
                "message": "没有找到活跃的RPD版本进行分析"
            }

        print(f"找到 {len(active_rpd_versions)} 个活跃的RPD版本")

        # 3. 并行生成findings（不保存到数据库）
        findings_data = await _generate_findings_for_image_analysis(
            image_bytes=image_bytes,
            image_width=image_width,
            image_height=image_height,
            active_rpd_versions=active_rpd_versions,
            mode=mode
        )

        # 4. 格式化结果
        formatted_findings = []
        for finding_data in findings_data:
            rpd_version = finding_data['rpd_version']
            finding_dict = finding_data['finding_dict']

            formatted_finding = {
                "rpd_id": str(rpd_version.review_point_definition_id),
                "rpd_key": rpd_version.parent_key,
                "rpd_title": rpd_version.title,
                "description": finding_dict['description'],
                "severity": finding_dict['severity'],
                "suggestion": finding_dict.get('suggestion', ''),
                "tag": finding_dict.get('tag', ''),
                "confidence": finding_dict.get('confidence', 0.8),  # 默认置信度
            }

            # 添加边界框信息（如果有）
            if 'bounding_boxes' in finding_data:
                formatted_finding['area'] = finding_data['bounding_boxes']

            formatted_findings.append(formatted_finding)

        processing_time = time.time() - start_time

        result = {
            "success": True,
            "findings": formatted_findings,
            "rpd_count": len(active_rpd_versions),
            "processing_time": processing_time,
            "image_info": {
                "filename": image_filename,
                "width": image_width,
                "height": image_height
            },
            "message": f"成功分析图片，发现 {len(formatted_findings)} 个问题点"
        }

        print(f"图片分析完成: {image_filename}, 处理时间: {processing_time:.2f}秒")
        return result

    except Exception as e:
        processing_time = time.time() - start_time
        error_message = f"图片分析失败: {str(e)}"
        print(error_message)

        return {
            "success": False,
            "findings": [],
            "rpd_count": 0,
            "processing_time": processing_time,
            "image_info": {
                "filename": image_filename,
                "width": 0,
                "height": 0
            },
            "error": error_message
        }


async def _generate_findings_for_image_analysis(
    image_bytes: bytes,
    image_width: int,
    image_height: int,
    active_rpd_versions: List[ReviewPointDefinitionVersionInDB],
    mode: AiReviewMode,
    ai_review_orm: Optional[AiReview] = None
) -> List[Dict[str, Any]]:
    """
    为图片分析生成findings，不保存到数据库
    这是_generate_findings_for_all_rpd_versions的简化版本
    """
    start_time = time.time()

    model_name = QUALITY_MODEL if mode == AiReviewMode.QUALITY else SPEED_MODEL
    logger.info(f"使用模型: {model_name} 进行图片分析")

    # 设置并发限制
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)  # 适中的并发数

    print(f"开始并行分析 {len(active_rpd_versions)} 个RPD版本")

    # 创建任务
    tasks = []
    for rpd_version in active_rpd_versions:
        task = asyncio.create_task(
            _generate_single_rpd_findings(
                image_bytes, image_width, image_height, rpd_version, semaphore, model_name
            ),
            name=f"rpd_analysis_{rpd_version.title}"
        )
        tasks.append(task)

    # 等待所有任务完成（如果提供了ai_review_orm，使用可中断机制）
    if ai_review_orm:
        done = await _wait_for_tasks_with_cancellation(tasks, ai_review_orm, 2.0)
        if not done:
            return []
    else:
        done, pending = await asyncio.wait(tasks, return_when=asyncio.ALL_COMPLETED)

    parallel_time = time.time() - start_time
    print(f"并行分析完成，耗时: {parallel_time:.2f}秒")

    # 收集结果
    all_findings_data = []
    for task in done:
        try:
            findings_data_list = await task
            if not isinstance(findings_data_list, Exception):
                all_findings_data.extend(findings_data_list)
        except Exception as e:
            print(f"任务执行异常: {e}")
            continue

    return all_findings_data


async def _get_image_dimensions(image_bytes: bytes) -> Tuple[int, int]:
    """
    获取图片尺寸
    """
    try:

        image = Image.open(io.BytesIO(image_bytes))
        return image.size  # (width, height)
    except Exception as e:
        print(f"获取图片尺寸失败: {e}")
        return (1920, 1080)  # 默认尺寸


async def get_latest_rpd_execution_summary(subtask_id: UUID) -> Dict:
    """获取最新AI审核的RPD执行摘要"""

    logger.info(f"正在查询subtask {subtask_id} 的最新AI审核记录...")

    # 首先检查是否有任何AI审核记录
    all_reviews = await AiReview.filter(subtask_id=subtask_id).all()
    logger.info(f"找到 {len(all_reviews)} 个AI审核记录")

    if all_reviews:
        for review in all_reviews:
            logger.info(
                f"AI审核记录: version={review.version}, is_latest={review.is_latest}, id={review.id}")

    # 获取最新的AI审核记录
    latest_review = await AiReview.filter(
        subtask_id=subtask_id,
        is_latest=True
    ).first()

    if latest_review:
        # 手动加载findings和相关数据
        await latest_review.fetch_related('findings__review_point_definition_version__review_point_definition')
        logger.info(f"加载findings后，数量: {len(latest_review.findings)}")

        # 如果还是没有findings，尝试直接查询
        if len(latest_review.findings) == 0:
            direct_findings = await AiReviewFindingEntry.filter(
                ai_review_id=latest_review.id
            ).prefetch_related('review_point_definition_version__review_point_definition').all()
            logger.info(f"直接查询findings数量: {len(direct_findings)}")

            # 如果直接查询有结果，使用直接查询的结果进行处理
            if direct_findings:
                findings_to_process = direct_findings
            else:
                findings_to_process = list(latest_review.findings)
        else:
            findings_to_process = list(latest_review.findings)

    if not latest_review:
        logger.info(f"未找到subtask {subtask_id} 的最新AI审核记录")
        return {
            'ai_review_version': None,
            'executed_at': None,
            'executed_rpds': [],
            'total_findings': 0
        }

    logger.info(
        f"找到最新AI审核记录: version={latest_review.version}, findings数量={len(findings_to_process)}")

    # 统计执行的RPD
    rpd_summary: Dict[str, Dict[str, Any]] = {}
    logger.info(f"开始处理 {len(findings_to_process)} 个findings...")

    for i, finding in enumerate(findings_to_process):
        logger.info(
            f"处理finding {i+1}: id={finding.id}, description={finding.description[:50]}...")

        try:
            # 检查是否正确预加载了关联数据
            rpd_version = finding.review_point_definition_version
            rpd = rpd_version.review_point_definition

            # 使用RPD ID作为唯一标识符，而不是key
            rpd_id = str(rpd.id)
            rpd_key = rpd.key

            logger.info(
                f"  RPD信息: id={rpd_id}, key={rpd_key}, title={rpd_version.title}, version={rpd_version.version_number}")

            if rpd_id not in rpd_summary:
                rpd_summary[rpd_id] = {
                    'rpd_key': rpd_key,
                    'rpd_title': rpd_version.title,
                    'version_number': rpd_version.version_number,
                    'finding_count': 0,
                    'severities': []
                }
            rpd_summary[rpd_id]['finding_count'] += 1
            rpd_summary[rpd_id]['severities'].append(finding.severity)

        except Exception as e:
            logger.error(f"  处理finding时出错: {e}")

            traceback.print_exc()

    logger.info(f"RPD统计结果: {list(rpd_summary.keys())}")
    logger.info(f"总findings数: {len(findings_to_process)}")

    return {
        'ai_review_version': latest_review.version,
        'executed_at': latest_review.review_timestamp,
        'executed_rpds': list(rpd_summary.values()),
        'total_findings': len(findings_to_process)
    }


# --- Single RPD Version Data Analysis (New Function) ---

async def analyze_single_image_with_rpd_data(
    image_bytes: bytes,
    image_filename: str,
    rpd_version_data: Dict[str, Any],
    cr_check: bool = False,
    mode: AiReviewMode = AiReviewMode.QUALITY,
    user_id: Optional[UUID] = None
) -> Dict[str, Any]:
    """
    使用传入的RPD version数据直接分析图片，不依赖数据库

    专为RPD测试设计，特点：
    - 接受RPD version的原始数据而不是数据库ID
    - 创建临时的RPD version对象用于分析
    - 不保存任何数据到数据库
    - 快速返回分析结果

    Args:
        image_bytes: 图片的字节数据
        image_filename: 图片文件名
        rpd_version_data: RPD version的数据字典，包含title, parent_key等字段
        cr_check: 是否启用版权检查（当前版本忽略此参数）
        mode: 分析模式（quality/speed）
        user_id: 用户ID（用于日志）

    Returns:
        Dict包含分析结果
    """
    start_time = time.time()

    try:
        print(f"开始使用RPD数据分析图片: {image_filename}")
        print(f"RPD数据: {rpd_version_data}")

        # 1. 获取图片尺寸
        image_width, image_height = await _get_image_dimensions(image_bytes)

        # 2. 创建临时的RPD version对象
        temp_rpd_version = await _create_temp_rpd_version_from_data(rpd_version_data)

        # 3. 生成findings（使用现有的单个RPD处理逻辑）
        model_name = QUALITY_MODEL if mode == AiReviewMode.QUALITY else SPEED_MODEL
        semaphore = asyncio.Semaphore(1)  # 单个RPD，不需要并发控制

        findings_data = await _generate_single_rpd_findings(
            image_bytes=image_bytes,
            image_width=image_width,
            image_height=image_height,
            rpd_version=temp_rpd_version,
            semaphore=semaphore,
            model_name=model_name
        )

        # 4. 格式化结果
        formatted_findings = []
        for finding_data in findings_data:
            finding_dict = finding_data['finding_dict']

            formatted_finding = {
                "rpd_key": temp_rpd_version.parent_key,
                "rpd_title": temp_rpd_version.title,
                "description": finding_dict['description'],
                "severity": finding_dict['severity'],
                "suggestion": finding_dict.get('suggestion', ''),
                "tag": finding_dict.get('tag', ''),
                "confidence": finding_dict.get('confidence', 0.8),  # 默认置信度
            }

            # 添加边界框信息（如果有）
            if 'bounding_boxes' in finding_data:
                formatted_finding['area'] = finding_data['bounding_boxes']

            formatted_findings.append(formatted_finding)

        processing_time = time.time() - start_time

        result = {
            "success": True,
            "findings": formatted_findings,
            "processing_time": processing_time,
            "image_info": {
                "filename": image_filename,
                "width": image_width,
                "height": image_height
            },
            "rpd_info": {
                "title": temp_rpd_version.title,
                "parent_key": temp_rpd_version.parent_key
            },
            "message": f"RPD测试完成，发现 {len(formatted_findings)} 个问题点"
        }

        print(f"RPD测试完成: {image_filename}, 处理时间: {processing_time:.2f}秒")
        return result

    except Exception as e:
        processing_time = time.time() - start_time
        error_message = f"RPD测试失败: {str(e)}"
        print(error_message)

        traceback.print_exc()

        return {
            "success": False,
            "findings": [],
            "processing_time": processing_time,
            "image_info": {
                "filename": image_filename,
                "width": 0,
                "height": 0
            },
            "error": error_message
        }


async def _create_temp_rpd_version_from_data(rpd_data: Dict[str, Any]) -> ReviewPointDefinitionVersionInDB:
    """
    从传入的数据创建临时的RPD version对象

    Args:
        rpd_data: 包含RPD version数据的字典

    Returns:
        ReviewPointDefinitionVersionInDB: 临时的RPD version对象
    """

    # 验证必需字段
    if not rpd_data.get('title'):
        raise ValueError("RPD title is required")
    if not rpd_data.get('parent_key'):
        raise ValueError("RPD parent_key is required")

    user_instruction = rpd_data.get('user_instruction', '')
    description_for_ai = user_instruction
    if rpd_data['parent_key'] == 'general_ng_review' and rpd_data.get('ng_subcategory'):
        if rpd_data.get('ng_subcategory') == 'abstract_type':
            prompt1, prompt2 = split_review_prompt(user_instruction)
            description_for_ai = prompt1+"\n********\n"+prompt2

    # 创建临时的RPD version对象
    temp_rpd_version = ReviewPointDefinitionVersionInDB(
        id=uuid.uuid4(),  # 临时ID
        review_point_definition_id=uuid.uuid4(),  # 临时ID
        title=rpd_data['title'],
        parent_key=rpd_data['parent_key'],
        user_instruction=user_instruction,
        description_for_ai=description_for_ai,
        ng_subcategory=rpd_data.get('ng_subcategory', None),
        tag_list=rpd_data.get('tag_list', []),
        reference_images=rpd_data.get('reference_images', []),
        reference_files=rpd_data.get('reference_files', []),
        special_rules=rpd_data.get('special_rules'),
        is_active_version=True,
        version_number=1,
        created_at=datetime.now(timezone.utc),
        created_by="test_user"
    )

    print(
        f"创建临时RPD version对象: {temp_rpd_version.title} ({temp_rpd_version.parent_key})")
    return temp_rpd_version


# ================================
# 其他内容类型处理函数（存根）
# ================================


async def _download_and_process_text(subtask_id: UUID, s3_path: str) -> Tuple[List[Dict[str, Any]], str]:
    """
    说明: 从 S3 下载并解析 Excel（.xlsx）对话稿，生成可并行处理的文本分段 `text_parts`。

    Input
    - subtask_id (UUID): 子任务 ID，用于日志或链路追踪。
    - s3_path (str): S3 文件路径，仅支持以 .xlsx 结尾的 Excel 文件。

    Output
    - Tuple[List[Dict[str, Any]], str]: 包含：
      - text_parts: 文本分段列表，每个元素包含：
        - original_index (int): 源 Excel 中的行索引（与 DataFrame 行号一致）
        - speaker (str): 「話者」列的值
        - line (str): 清洗后的正文（去除 ★、※、#、换行、全角空格）
        - sheet_index (int): 工作表序号，从 0 开始
      - local_file_path: 本地临时Excel文件路径

    约束/行为: 
      - 每个工作表 `skiprows=11`, `engine='openpyxl'`
      - 正文列名固定为 "本文３４５６７８９０１２３４５６７８９０１２３４５６７８９０１２３４５６"
      - 无法读取的工作表将被跳过
    - 若 `s3_path` 非 `.xlsx` 将抛出 `ValueError`
    """
    if not s3_path or not s3_path.lower().endswith(".xlsx"):
        raise ValueError("只支持处理 .xlsx 文件")

    # 下载 S3 文件内容到临时文件
    excel_bytes = download_file_content_from_s3_sync(s3_path)

    # 创建临时文件
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        tmp_file.write(excel_bytes)
        local_file_path = tmp_file.name

    # 从本地文件创建buffer用于pandas读取
    bytes_buffer = io.BytesIO(excel_bytes)

    # 与 batch.speaker_check_process_s3_file 对齐的列名
    text_col = "本文３４５６７８９０１２３４５６７８９０１２３４５６７８９０１２３４５６"

    try:
        xls = pd.ExcelFile(bytes_buffer)
    except Exception as exc:
        raise ValueError(f"读取Excel失败: {exc}")

    text_parts: List[Dict[str, Any]] = []

    for sheet_index, sheet_name in enumerate(xls.sheet_names):
        try:
            df = pd.read_excel(
                xls,
                sheet_name=sheet_name,
                skiprows=11,
                engine="openpyxl",
            )
        except Exception:
            # 跳过无法读取的工作表
            continue

        for row_index, row in df.iterrows():
            speaker = row.get("話者")
            line = row.get(text_col)
            if pd.notna(speaker) and pd.notna(line):
                clean_line = (
                    str(line)
                    .replace("★", "")
                    .replace("※", "")
                    .replace("#", "")
                    .replace("\n", "")
                    .replace("\u3000", "")
                )
                text_parts.append(
                    {
                        "original_index": row_index,
                        "speaker": speaker,
                        "line": clean_line,
                        "sheet_index": sheet_index,
                    }
                )
    return text_parts, local_file_path


async def _download_and_process_audio(subtask_id: UUID, s3_path: str) -> tuple[bytes, dict]:
    """下载并处理音频，返回音频字节数据和元数据

    Args:
        subtask_id: 子任务ID
        s3_path: S3路径

    Returns:
        tuple: (音频字节数据, 音频元数据字典)

    TODO: 实现音频处理逻辑
    - 下载音频文件
    - 提取音频特征
    - 获取音频元数据（时长、格式等）
    - 可能需要语音转文本
    """
    raise NotImplementedError("音频处理功能尚未实现")


# async def _split_audio_into_segments(audio_data: bytes, audio_metadata: dict, segment_duration_seconds: int = 30) -> List[Dict[str, Any]]:
#     """
#     将长音频分割成多个片段以便并行处理

#     Args:
#         audio_data: 完整的音频字节数据
#         audio_metadata: 音频元数据（包含时长、采样率等信息）
#         segment_duration_seconds: 每个片段的时长（秒）

#     Returns:
#         List[Dict[str, Any]]: 音频片段列表，每个元素包含 {'data': bytes, 'metadata': dict}

#     TODO: 实现智能音频分割逻辑
#     - 按固定时长分割
#     - 按静音检测分割
#     - 按语音段落分割
#     - 保持音频质量
#     """
#     if not audio_data:
#         return []

#     audio_segments = []
#     total_duration = audio_metadata.get('duration_seconds', 0)

#     if total_duration == 0:
#         print("音频时长信息缺失，无法分割")
#         return []

#     # 简单的基于时长的分割实现（实际应该使用音频处理库如librosa、pydub等）
#     # 实际实现中应该考虑采样率、声道数等音频参数
#     bytes_per_second = len(audio_data) / \
#         total_duration if total_duration > 0 else 0

#     for i in range(0, int(total_duration), segment_duration_seconds):
#         start_time = i
#         end_time = min(i + segment_duration_seconds, total_duration)

#         # 计算字节范围（这是简化的计算，实际需要考虑音频格式）
#         start_byte = int(start_time * bytes_per_second)
#         end_byte = int(end_time * bytes_per_second)

#         segment_data = audio_data[start_byte:end_byte]
#         segment_metadata = {
#             'segment_number': len(audio_segments) + 1,
#             'start_time': start_time,
#             'end_time': end_time,
#             'duration': end_time - start_time,
#             'total_segments': (int(total_duration) + segment_duration_seconds - 1) // segment_duration_seconds,
#             'original_metadata': audio_metadata
#         }

#         audio_segments.append({
#             'data': segment_data,
#             'metadata': segment_metadata
#         })

#     print(f"音频已分割为 {len(audio_segments)} 个片段")
#     return audio_segments


# async def _generate_findings_for_video(
#     video_data: bytes,
#     video_metadata: dict,
#     active_rpd_versions: List,
#     ai_review_orm,
#     mode: AiReviewMode
# ) -> List:
#     """为视频内容生成AI审查结果

#     Args:
#         video_data: 视频字节数据
#         video_metadata: 视频元数据
#         active_rpd_versions: 活跃的审查点定义版本列表
#         ai_review_orm: AI审查ORM对象
#         mode: 审查模式

#     Returns:
#         List: AI审查发现条目列表

#     TODO: 实现视频AI审查逻辑
#     - 对关键帧进行图像审查
#     - 检测视频中的违规内容
#     - 分析视频场景和内容
#     """
#     raise NotImplementedError("视频AI审查功能尚未实现")


async def _generate_findings_for_text(
    text_content: str,
    text_metadata: dict,
    active_rpd_versions: List,
    ai_review_orm,
    mode: AiReviewMode
) -> List:
    """为文本内容生成AI审查结果

    Args:
        text_content: 文本内容
        text_metadata: 文本元数据
        active_rpd_versions: 活跃的审查点定义版本列表
        ai_review_orm: AI审查ORM对象
        mode: 审查模式

    Returns:
        List: AI审查发现条目列表

    TODO: 实现文本AI审查逻辑
    - 检测文本中的违规内容
    - 分析文本情感和内容
    - 检查文本格式和质量
    """
    raise NotImplementedError("文本AI审查功能尚未实现")


# async def _generate_findings_for_audio(
#     audio_data: bytes,
#     audio_metadata: dict,
#     active_rpd_versions: List,
#     ai_review_orm,
#     mode: AiReviewMode
# ) -> List:
#     """为音频内容生成AI审查结果

#     Args:
#         audio_data: 音频字节数据
#         audio_metadata: 音频元数据
#         active_rpd_versions: 活跃的审查点定义版本列表
#         ai_review_orm: AI审查ORM对象
#         mode: 审查模式

#     Returns:
#         List: AI审查发现条目列表

#     TODO: 实现音频AI审查逻辑
#     - 语音转文本后进行文本审查
#     - 检测音频中的违规内容
#     - 分析音频质量和内容
#     """
#     raise NotImplementedError("音频AI审查功能尚未实现")


# async def _process_single_audio_segment(
#     audio_segment: bytes,
#     audio_metadata: Dict[str, Any],
#     content_type: SubtaskType,
#     active_rpd_versions: List[ReviewPointDefinitionVersionInDB],
#     ai_review_orm: AiReview,
#     mode: AiReviewMode,
#     semaphore: asyncio.Semaphore
# ) -> List[AiReviewFindingEntry]:
#     """
#     处理单个音频片段，为其生成所有RPD的findings
#     """
#     async with semaphore:
#         try:
#             print(
#                 f"处理音频片段 - {audio_metadata.get('segment_number', 'unknown')}")

#             # 这里需要实现音频的AI审查逻辑
#             # 目前先返回空列表，等待音频处理功能的完整实现
#             findings = await _generate_findings_for_audio(
#                 audio_segment, audio_metadata, active_rpd_versions, ai_review_orm, mode
#             )

#             print(
#                 f"音频片段 {audio_metadata.get('segment_number', 'unknown')} 处理完成")
#             return findings

#         except NotImplementedError:
#             print(
#                 f"音频处理功能尚未实现，跳过音频片段 {audio_metadata.get('segment_number', 'unknown')}")
#             return []
#         except Exception as e:
#             print(f"处理音频片段时出错: {e}")
#             return []


# async def _generate_findings_for_audio_segments_parallel(
#     # 每个元素包含 {'data': bytes, 'metadata': dict}
#     audio_segments: List[Dict[str, Any]],
#     content_type: SubtaskType,
#     active_rpd_versions: List[ReviewPointDefinitionVersionInDB],
#     ai_review_orm: AiReview,
#     mode: AiReviewMode,
# ) -> List[AiReviewFindingEntry]:
#     """
#     并行处理所有音频片段，为每个片段生成AI审查findings
#     """
#     start_time = time.time()

#     print(f"开始并行处理 {len(audio_segments)} 个音频片段")

#     # 设置并发信号量
#     semaphore = asyncio.Semaphore(max_concurrent)

#     # 创建所有音频片段处理任务
#     tasks = []
#     for i, audio_segment_data in enumerate(audio_segments):
#         audio_data = audio_segment_data['data']
#         audio_metadata = audio_segment_data['metadata']

#         task = asyncio.create_task(
#             _process_single_audio_segment(
#                 audio_data, audio_metadata, content_type,
#                 active_rpd_versions, ai_review_orm, mode, semaphore
#             ),
#             name=f"audio_segment_task_{i}"
#         )
#         tasks.append(task)

#     print(f"所有 {len(tasks)} 个音频片段任务已创建，开始等待完成...")

#     # 等待所有任务完成
#     done, pending = await asyncio.wait(tasks, return_when=asyncio.ALL_COMPLETED)

#     # 收集所有结果
#     all_findings = []
#     for task in done:
#         try:
#             findings = await task
#             all_findings.extend(findings)
#         except Exception as e:
#             print(f"获取音频片段处理结果时出错: {e}")

#     end_time = time.time()
#     print(f"音频片段并行处理完成，总耗时: {end_time - start_time:.2f}秒")
#     print(f"共生成 {len(all_findings)} 个findings")

#     return all_findings

    # except Exception as db_error:
    #     logger.error(f"🗄️ [批次 {batch_id}] 更新数据库记录失败: {str(db_error)}")
    #     # 即使数据库更新失败，也不影响主要任务的完成


async def interrupt_ai_review_for_subtask(subtask_id: UUID) -> bool:
    """
    中断指定子任务的AI审查处理。

    Args:
        subtask_id: 子任务ID

    Returns:
        bool: 如果成功中断返回True，如果没有找到可中断的review返回False
    """
    try:
        # 获取最新的AI review
        latest_review = await get_latest_ai_review_for_subtask(subtask_id)

        if not latest_review:
            logger.info(f"No AI review found for subtask {subtask_id}")
            return False

        # 检查是否可以中断（只有processing状态的可以中断）
        if latest_review.processing_status != AiReviewProcessingStatus.PROCESSING.value:
            logger.info(
                f"AI review for subtask {subtask_id} is not in processing state, current state: {latest_review.processing_status}")
            return False

        # 获取AiReview ORM对象并更新状态为cancelled
        ai_review_orm = await AiReview.get(id=latest_review.id).prefetch_related('subtask')

        # 首先设置中断信号，让正在运行的任务能够检测到
        logger.info(
            f"Setting cancellation signal for AI review {latest_review.id}")
        ai_review_orm.should_cancel = True
        await ai_review_orm.save(update_fields=['should_cancel'])

        # 然后更新处理状态
        logger.info(
            f"Updating processing status to CANCELLED for AI review {latest_review.id}")
        await _update_processing_status(
            ai_review_orm,
            AiReviewProcessingStatus.CANCELLED,
            error_message="AI review was interrupted by user"
        )

        # 执行回退逻辑
        logger.info(
            f"AI review cancelled for subtask {subtask_id}, attempting rollback...")
        await _rollback_failed_review(ai_review_orm)

        logger.info(
            f"Successfully interrupted AI review {latest_review.id} for subtask {subtask_id}")
        return True

    except Exception as e:
        logger.error(
            f"Error interrupting AI review for subtask {subtask_id}: {e}")
        traceback.print_exc()
        return False
